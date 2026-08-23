"""
Turn the SAP exports the courier delivers into work-planning candidates.

Reads IW39 (the orders) joined to IW49 (the operations, which carry the hours),
and reports what it found. Deliberately does NOT write to the pool yet — the
matched/unmatched report has to be trustworthy before anything is created from it.

Every rule below was verified against Ali's real 19,283-row IW39 export, and each
one exists because the obvious alternative was silently wrong:

  * MES only. `Main work center` starting with MES is Ali's section; CMS is the
    crane side and is not his. Without this the pool would carry another team's work.

  * `MaintActivityType`, NOT `Order Type`. Order Type holds SDMN/USMN/SBOR/...;
    PRM/COM/INS live in MaintActivityType. Reading the wrong column maps every
    order to 'pm' through the importer's quiet fallback.

  * Equipment from `Functional Location`, NOT `Equipment`. The Equipment column
    is empty on 74% of rows and otherwise holds SAP's internal id
    (800000000010092), which matches nothing in the app. Functional Location is
    populated on 19,280/19,283 and carries the plant code the app actually uses
    (3700-EQ-ECH-ECH02 -> ECH02).

  * Hours converted per row. `Unit for work` is a MIX of H, HR and MIN — 28% of
    the 56,131 operation rows are MINUTES. Summing blindly turns a 30-minute
    operation into 30 hours.

  * Order numbers read as strings. pandas otherwise renders 700001479825 as
    7.000015e+11.
"""

import io
import logging
import re
from collections import defaultdict

logger = logging.getLogger(__name__)

# Ali's section. CMS (cranes), FMS and REF belong to other teams.
MES_PREFIX = 'MES'

# SAP's `System status` is a compound string — "CRTD CSER NMAT PRC" or
# "TECO CNF CSER NMAT PRC SETC" — so these are matched as TOKENS, never as
# substrings or equality.
#
# Ali's rule: an order may be planned only while it is created or released, and
# never once it is confirmed, technically complete, or closed.
#
# CNF matters on its own. An order can be fully confirmed without yet being
# TECO'd, and a "not TECO and not CLSD" test would let that through — the work is
# already done, so planning it again would send a crew to a finished job.
PLANNABLE_STATUS_TOKENS = ('CRTD', 'REL')
BLOCKING_STATUS_TOKENS = ('CLSD', 'TECO', 'CNF')

# Confirmed by Ali. Everything else — notably BDM (breakdown, 5,149 rows YTD) —
# is handled reactively and is deliberately not planned.
PLANNABLE_ACTIVITY_TYPES = ('PRM', 'COM', 'DAM', 'INS', 'ACD')

# MaintActivityType -> the app's job_type CHECK constraint
# (pm | defect | inspection | corrective).
# DAM (damage) and ACD map to 'corrective' as a considered default — worth
# confirming with Ali once real jobs are flowing; only 2 of 208 open orders today.
ACTIVITY_TO_JOB_TYPE = {
    'PRM': 'pm',
    'COM': 'defect',
    'INS': 'inspection',
    'DAM': 'corrective',
    'ACD': 'corrective',
}

# SAP priority is numeric 1..4, 1 being most urgent.
SAP_PRIORITY_TO_APP = {'1': 'urgent', '2': 'high', '3': 'normal', '4': 'low'}

# 3700-EQ-ECH-ECH02-SUBPART  ->  ECH02
FUNCTIONAL_LOCATION_RE = re.compile(r'^\d+-[A-Z]{2}-[A-Z_]{3}-([A-Z]{2,4}\d{2,4})')

# Rows whose unit is blank are treated as hours (the majority unit) and counted
# separately so the assumption stays visible rather than buried.
MINUTE_UNITS = {'MIN'}
HOUR_UNITS = {'H', 'HR', 'HRS', 'STD'}

DEFAULT_HOURS = 4.0



def is_plannable_status(system_status):
    """Whether an order may enter the job pool, from its SAP system status.

    Ali's rule: must carry CRTD or REL, and must not carry CLSD, TECO or CNF.
    Tokenised on whitespace — 'REL' must not match inside 'RELEASED_X', and
    'CNF' must not match a substring of some other code.
    """
    if system_status is None or isinstance(system_status, float):
        return False
    tokens = set(str(system_status).upper().split())
    if tokens & set(BLOCKING_STATUS_TOKENS):
        return False
    return bool(tokens & set(PLANNABLE_STATUS_TOKENS))


# SAP marks a cancelled order with the user status CNCL. Verified on the real
# export: 2 of 9,124 MES orders carry it, and BOTH also carry CLSD in the system
# status — so cancellation must be tested FIRST or a cancelled order is
# indistinguishable from a finished one.
CANCELLED_USER_STATUS_TOKENS = ('CNCL',)


def order_sap_state(system_status, user_status=None, deletion_flag=None):
    """What SAP now says about one order: cancelled, done, open or unknown.

    The inverse of is_plannable_status(). That answers "may this enter the pool";
    this answers "what happened to the one already on Tuesday's plan".

    'unknown' is a real answer and deliberately triggers nothing. An order can be
    absent from an export, carry a status nobody has seen before, or belong to a
    work centre outside MES. Acting on any of those would mean treating silence
    as evidence, and the whole point of these rules is that a job on someone's
    day never vanishes without a reason we can name.
    """
    flag = '' if (deletion_flag is None or isinstance(deletion_flag, float)) else str(deletion_flag).strip()
    user_tokens = set() if (user_status is None or isinstance(user_status, float)) \
        else set(str(user_status).upper().split())
    if flag or (user_tokens & set(CANCELLED_USER_STATUS_TOKENS)):
        return 'cancelled'

    if system_status is None or isinstance(system_status, float):
        return 'unknown'
    tokens = set(str(system_status).upper().split())
    if tokens & set(BLOCKING_STATUS_TOKENS):
        return 'done'
    if tokens & set(PLANNABLE_STATUS_TOKENS):
        return 'open'
    return 'unknown'


def build_order_status_index(iw39_bytes):
    """order number -> what SAP says about it now.

    Covers EVERY MES order in the export, open or closed — unlike
    parse_open_orders, which keeps only the plannable ones. Reconciliation needs
    the closed rows: they are the evidence that a scheduled job is finished.
    """
    import pandas as pd

    wanted = ['Order', 'Main work center', 'System status', 'User Status',
              'Deletion flag', 'Actual Order Finish Date']
    df = _read_excel(iw39_bytes, wanted)
    mes = df[df['Main work center'].fillna('').str.upper().str.startswith(MES_PREFIX)]

    finish_dates = pd.to_datetime(mes['Actual Order Finish Date'], errors='coerce')

    index = {}
    for order, system_status, user_status, flag, finished in zip(
            mes['Order'], mes['System status'], mes['User Status'],
            mes['Deletion flag'], finish_dates):
        if order is None or isinstance(order, float):
            continue
        number = str(order).strip()
        if not number:
            continue
        index[number] = {
            'state': order_sap_state(system_status, user_status, flag),
            'system_status': None if isinstance(system_status, float) else system_status,
            'finished_on': None if pd.isna(finished) else finished.date(),
        }
    return index



def _today_naive(today=None):
    """A timezone-NAIVE timestamp for today in the yard's timezone.

    Naive on purpose: the dates parsed out of SAP exports are naive, and
    pandas refuses to compare naive against aware. pd.Timestamp.utcnow() is
    AWARE, so using it here raised "Cannot compare tz-naive and tz-aware" —
    but only when no explicit `today` was passed, i.e. only in production.

    Baghdad rather than UTC for the same reason planning_today() exists: at
    UTC+3 the server believes it is still yesterday until 03:00 local, which
    silently shifts every window by a day.
    """
    import pandas as pd
    if today is not None:
        return pd.Timestamp(today)
    from datetime import datetime, timedelta, timezone
    return pd.Timestamp((datetime.now(timezone.utc) + timedelta(hours=3)).date())


def _read_excel(source, usecols):
    """Parse an export, or reuse an already-parsed one.

    Accepts raw bytes OR a DataFrame. IW39 is consumed by four separate
    functions (last completion, breakdowns, durations, and the main parse), and
    re-reading a 10 MB / 144-column workbook four times dominated the runtime.
    Callers that need it more than once parse it ONCE with the union of columns
    and pass the frame.

    Reading the union rather than caching whole workbooks is deliberate: IW39
    holds 19,283 rows x 144 columns, and holding all of that as object dtype
    would be hundreds of MB on a 512 MB instance. Fifteen columns is nothing.
    """
    import pandas as pd

    if isinstance(source, pd.DataFrame):
        missing = [c for c in usecols if c not in source.columns]
        if missing:
            raise KeyError(f'pre-parsed frame is missing columns: {missing}')
        return source[list(usecols)]
    return _stream_excel(source, usecols)


def _cell_to_str(value):
    """Match what pandas `dtype=str` produces, cell for cell.

    The streaming reader below returns native types where read_excel returned
    strings, and the rest of this module was written against strings — dates
    arrive as '2025-08-31 00:00:00' and are re-parsed with pd.to_datetime,
    order numbers as '700001479825'. Converting here keeps every downstream
    rule untouched instead of rewriting nine functions around a new dtype.
    """
    if value is None:
        return None
    if isinstance(value, str):
        # An empty cell reaches openpyxl as '' but reaches pandas as NaN, and
        # the whole module is written around NaN meaning "not filled in".
        # Returning '' here would make .fillna() and isinstance(x, float)
        # checks quietly stop working on exactly the blank cells they guard.
        return value if value else None
    return str(value)


def _stream_excel(raw_bytes, usecols):
    """Read only the wanted columns, one row at a time.

    pd.read_excel(usecols=...) does NOT save memory: openpyxl still builds the
    whole worksheet before pandas selects columns. Measured on the real exports,
    the three files peaked at 647 MB — on a 512 MB container, with two gunicorn
    workers already resident. The parse was killed outright, leaving no
    traceback, which is why it looked like the job simply never finished.

    read_only=True makes openpyxl stream rows instead of building an object
    tree, so peak memory becomes the size of the columns actually kept rather
    than the size of the workbook.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        rows = workbook[workbook.sheetnames[0]].iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError('worksheet is empty')
        return rows_to_frame(header, rows, usecols)
    finally:
        workbook.close()


def rows_to_frame(header, rows, usecols):
    """Build the frame from a header row and an iterator of row tuples.

    Separate from the workbook so it can be tested with plain tuples. Both bugs
    that showed up here — the phantom trailing row and empty-string cells — are
    invisible through a written fixture, because both openpyxl and pandas
    normalise '' to None on the way OUT to a file. They only exist coming IN
    from SAP's exports.
    """
    import pandas as pd

    position = {}
    for index, name in enumerate(header):
        if name is not None and name not in position:
            position[name] = index

    missing = [c for c in usecols if c not in position]
    if missing:
        # Loud on purpose. A silently missing column would disable a rule
        # (cancellation detection, hours, the plant code) while every number
        # downstream still looked plausible.
        raise KeyError(f'export is missing columns: {missing}')

    wanted = [position[c] for c in usecols]
    width = max(wanted) + 1
    columns = [[] for _ in usecols]

    # A worksheet's declared dimensions routinely over-report the used range, so
    # streaming yields trailing rows that hold nothing. read_excel drops them,
    # and the row counts must match or every check against the old behaviour is
    # off by one. Only TRAILING blanks go: a blank row in the middle is real
    # data shaped like a gap, and read_excel keeps it.
    last_populated = -1
    for count, row in enumerate(rows):
        if len(row) < width:
            row = tuple(row) + (None,) * (width - len(row))
        for slot, index in enumerate(wanted):
            columns[slot].append(_cell_to_str(row[index]))
        # '' counts as blank, not content. IK17's final row is 85 empty STRINGS
        # rather than 85 Nones, and an `is not None` test kept it — one phantom
        # row, enough to put every row count out by one.
        if any(value is not None and value != '' for value in row):
            last_populated = count

    keep = last_populated + 1
    if keep < len(columns[0]):
        columns = [values[:keep] for values in columns]

    return pd.DataFrame({name: values for name, values in zip(usecols, columns)},
                        columns=list(usecols))


# Every IW39 column any consumer in this module needs. Read once, shared.
IW39_COLUMNS = (
    'Order', 'MaintActivityType', 'Main work center', 'System status',
    'Functional Location', 'Description', 'Basic start date', 'Priority',
    'Work Center', 'Equipment', 'Maintenance Plan', 'Created on',
    'Actual Order Finish Date', 'User Status', 'Deletion flag',
)


def load_iw39(iw39_bytes):
    """Parse IW39 once, with every column this module needs."""
    return _read_excel(iw39_bytes, list(IW39_COLUMNS))


def parse_operation_hours(iw49_bytes):
    """Planned hours per order, summed across operations and unit-corrected.

    IW39 carries no hours column at all, so this is the only source. Returns
    ({order_number: hours}, report).
    """
    import pandas as pd

    wanted = ['Order', 'Work', 'Unit for work']
    df = _read_excel(iw49_bytes, wanted)

    hours = defaultdict(float)
    counts = {'rows': len(df), 'minutes': 0, 'hours': 0, 'unitless': 0, 'unparseable': 0}

    values = pd.to_numeric(df['Work'], errors='coerce')
    for order, value, unit in zip(df['Order'], values, df['Unit for work']):
        if not order or (isinstance(order, float)) or pd.isna(value):
            counts['unparseable'] += 1
            continue
        # NaN arrives as a float when the cell is blank, so coerce before stripping.
        u = '' if (unit is None or isinstance(unit, float)) else str(unit).strip().upper()
        if u in MINUTE_UNITS:
            hours[str(order).strip()] += float(value) / 60.0
            counts['minutes'] += 1
        elif u in HOUR_UNITS:
            hours[str(order).strip()] += float(value)
            counts['hours'] += 1
        else:
            # Unknown or blank unit: assume hours, but say so.
            hours[str(order).strip()] += float(value)
            counts['unitless'] += 1

    return dict(hours), counts


def extract_plant_code(functional_location):
    """The equipment code the app knows, pulled out of the functional location."""
    if not functional_location:
        return None
    m = FUNCTIONAL_LOCATION_RE.match(str(functional_location).strip().upper())
    return m.group(1) if m else None


def parse_open_orders(iw39_bytes, hours_by_order=None, plan_types=None,
                     last_completion=None, today=None):
    """Open, plannable MES orders from an IW39 export.

    Returns (candidates, report). Nothing is written; the caller decides.
    """
    import pandas as pd

    wanted = ['Order', 'MaintActivityType', 'Main work center', 'System status',
              'Functional Location', 'Description', 'Basic start date', 'Priority',
              'Work Center', 'Equipment', 'Maintenance Plan', 'Created on']
    df = _read_excel(iw39_bytes, wanted)

    plan_types = plan_types or {}
    last_completion = last_completion or {}
    now = _today_naive(today)

    report = {'total_rows': len(df)}

    mes = df[df['Main work center'].fillna('').str.upper().str.startswith(MES_PREFIX)]
    report['mes_rows'] = len(mes)

    open_orders = mes[mes['System status'].map(is_plannable_status)]
    report['open_rows'] = len(open_orders)

    # An order whose MaintActivityType is blank but whose maintenance plan Ali has
    # classified is still real work — SAP leaves that field empty occasionally
    # (e.g. 700001778121, TT032-1000HR-MECH. HOURLY SERVICE). Discarding it
    # because of a missing field would silently drop a PM that is genuinely due.
    activity_ok = open_orders['MaintActivityType'].isin(PLANNABLE_ACTIVITY_TYPES)
    plan_is_classified = (
        open_orders['MaintActivityType'].isna()
        & open_orders['Maintenance Plan'].fillna('').str.strip().isin(plan_types.keys())
    )
    plannable = open_orders[activity_ok | plan_is_classified]
    report['plannable_rows'] = len(plannable)
    report['excluded_activity_types'] = (
        open_orders[~open_orders['MaintActivityType'].isin(PLANNABLE_ACTIVITY_TYPES)]
        ['MaintActivityType'].value_counts(dropna=False).to_dict()
    )

    hours_by_order = hours_by_order or {}
    candidates = []
    unparsed_locations = []

    def cell(record, column):
        """Read a column by its real name.

        Deliberately not itertuples(): several of these columns contain spaces
        ('Functional Location', 'Basic start date'), which itertuples renames to
        positional _1/_2/... — and the positions follow the file's column order,
        not the order requested in usecols. Reading by name cannot drift.
        """
        value = record.get(column)
        if value is None or (isinstance(value, float)):
            return None
        text = str(value).strip()
        return text or None

    for record in plannable.to_dict('records'):
        order_number = cell(record, 'Order')
        if not order_number:
            continue

        functional_location = cell(record, 'Functional Location')
        plant_code = extract_plant_code(functional_location)
        if not plant_code:
            unparsed_locations.append(functional_location)

        activity = (cell(record, 'MaintActivityType') or '').upper()

        maintenance_plan = cell(record, 'Maintenance Plan')
        pm_basis = plan_types.get(maintenance_plan) if maintenance_plan else None

        created_on = pd.to_datetime(cell(record, 'Created on'), errors='coerce')
        age_days = None if pd.isna(created_on) else (now - created_on).days

        last_done = last_completion.get(maintenance_plan) if maintenance_plan else None
        days_since_last_pm = None if last_done is None else (now - last_done).days

        if pm_basis == 'calendar':
            priority = calendar_pm_priority(age_days, days_since_last_pm)
        else:
            # Hourly PMs and all non-PM work keep a neutral priority until their
            # own rule is agreed. SAP's Priority field is deliberately unused —
            # see the note above calendar_pm_priority.
            priority = 'normal'

        candidates.append({
            'order_number': order_number,
            'activity_type': activity,
            # A classified maintenance plan means PM even when SAP left the
            # activity type blank.
            'job_type': ACTIVITY_TO_JOB_TYPE.get(activity, 'pm'),
            'plant_code': plant_code,
            'functional_location': functional_location,
            'description': cell(record, 'Description'),
            'required_date': cell(record, 'Basic start date'),
            'priority': priority,
            'maintenance_plan': maintenance_plan,
            'pm_basis': pm_basis,
            # Carried through so the corrective rule can see it. REL means a
            # planner released the order for work; CRTD means it is only created.
            'system_status': cell(record, 'System status'),
            'is_released': 'REL' in set((cell(record, 'System status') or '').upper().split()),
            'age_days': age_days,
            'days_since_last_pm': days_since_last_pm,
            'work_center': cell(record, 'Work Center'),
            'estimated_hours': round(hours_by_order.get(order_number, DEFAULT_HOURS), 2),
            'hours_from_iw49': order_number in hours_by_order,
        })

    report['candidates'] = len(candidates)
    report['unparsed_functional_locations'] = len(unparsed_locations)
    report['sample_unparsed'] = [str(x) for x in unparsed_locations[:5]]
    report['without_iw49_hours'] = sum(1 for c in candidates if not c['hours_from_iw49'])
    report['unclassified_plans'] = sorted({c['maintenance_plan'] for c in candidates
                                           if c['job_type'] == 'pm' and not c['pm_basis']
                                           and c['maintenance_plan']})
    report['by_pm_basis'] = {}
    report['by_priority'] = {}
    for c in candidates:
        report['by_pm_basis'][c['pm_basis'] or 'n/a'] = report['by_pm_basis'].get(c['pm_basis'] or 'n/a', 0) + 1
        report['by_priority'][c['priority']] = report['by_priority'].get(c['priority'], 0) + 1

    return candidates, report


def match_equipment(candidates):
    """Resolve plant codes against the app's Equipment table.

    The one failure in this pipeline that would otherwise be SILENT: unmatched
    equipment means the order is dropped and the planner simply looks empty, with
    nothing explaining why. So every unmatched code is named in the report.
    """
    from app.models import Equipment

    codes = {c['plant_code'] for c in candidates if c['plant_code']}
    if not codes:
        return {'matched': 0, 'unmatched': 0, 'unmatched_codes': [], 'match_rate': 0.0}

    rows = Equipment.query.filter(
        (Equipment.serial_number.in_(codes)) | (Equipment.name.in_(codes))
    ).all()

    known = set()
    lookup = {}
    for eq in rows:
        for key in (eq.serial_number, eq.name):
            if key in codes:
                known.add(key)
                lookup[key] = eq.id

    for c in candidates:
        c['equipment_id'] = lookup.get(c['plant_code'])

    unmatched = sorted(codes - known)
    return {
        'matched': len(known),
        'unmatched': len(unmatched),
        'unmatched_codes': unmatched,
        'match_rate': round(len(known) / len(codes) * 100, 1),
        'distinct_codes': len(codes),
    }

# ---------------------------------------------------------------------------
# PM priority — Ali's rules, 2026-08-22
# ---------------------------------------------------------------------------
#
# SAP's own Priority field is deliberately ignored. On the open backlog it does
# not indicate urgency: 51% of open orders are priority 1 versus 2.3% of
# finished ones, and every one of those is over 90 days old. It marks stale
# corrective orders, not work that must happen first.
#
# Calendar PMs are judged on TWO signals, whichever is worse:
#   1. How long the order has been sitting     (today - Created on)
#   2. How long since this SERVICE last finished (today - last completion of the
#      same maintenance plan)
#
# The second exists because the first can lie. A freshly-raised order looks
# harmless at 7 days old even when that inspection has not been done for 376.
# Real example: TT031 "Inspection AC System" — order age 7, last completed 376
# days ago. Signal 1 says "high", signal 2 says urgent. Signal 2 is right.
#
# Matching is on MAINTENANCE PLAN, not equipment — "when was THIS service last
# done", not "when did this machine last get any attention". Ali's choice, and
# the stricter of the two.

CALENDAR_URGENT_AGE_DAYS = 10
CALENDAR_HIGH_AGE_DAYS = 5
# Two months without this service is urgent regardless of how new the order is.
CALENDAR_URGENT_SINCE_LAST_PM_DAYS = 60


def calendar_pm_priority(age_days, days_since_last_pm):
    """Priority for a calendar-based PM.

    Returns one of the app's four values (low/normal/high/urgent) — note there
    is no 'medium'; Ali's "medium" tier is the app's existing 'high'.
    """
    if days_since_last_pm is not None and days_since_last_pm > CALENDAR_URGENT_SINCE_LAST_PM_DAYS:
        return 'urgent'
    if age_days is None:
        # No creation date to judge by, and no overdue service — don't invent urgency.
        return 'normal'
    if age_days > CALENDAR_URGENT_AGE_DAYS:
        return 'urgent'
    if age_days > CALENDAR_HIGH_AGE_DAYS:
        return 'high'
    return 'normal'


def load_maintenance_plan_types(xlsx_bytes):
    """Ali's hand-maintained hourly/calendar classification.

    SAP does not carry the plan's basis in IW39 — only in the plan definition
    (IP24), which is not exported. Deriving it from order rhythm was tried and
    failed: calendar plans fire with a CV of 1.10, i.e. no rhythm at all, because
    execution is irregular. So this file is the authority.

    Sheet names and column names differ between the two tabs ('M.P' vs
    'MaintenancePlan'), so both spellings are accepted rather than requiring Ali
    to reformat a file he maintains by hand.
    """
    import pandas as pd

    types = {}
    workbook = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    for sheet in workbook.sheet_names:
        frame = pd.read_excel(workbook, sheet_name=sheet, dtype=str)
        plan_col = next((c for c in frame.columns
                         if str(c).strip().lower().replace('.', '').replace(' ', '')
                         in ('mp', 'maintenanceplan', 'plan')), None)
        type_col = next((c for c in frame.columns
                         if str(c).strip().lower() == 'type'), None)
        if not plan_col or not type_col:
            logger.warning('maintenance plan sheet %r has no recognisable columns (%s)',
                           sheet, list(frame.columns))
            continue
        for plan, kind in zip(frame[plan_col], frame[type_col]):
            if plan is None or isinstance(plan, float):
                continue
            key = str(plan).strip()
            value = str(kind).strip().lower() if kind is not None else ''
            if key and value in ('hourly', 'calendar'):
                types[key] = value
    return types


def build_last_completion_index(iw39_bytes):
    """Most recent actual finish date per maintenance plan.

    Reads the FINISHED orders — the ones filtered out of the pool — because the
    question is "when was this service last actually done".

    Trap: 'Actual finish' is a TIME column ('15:00:00'). The date lives in
    'Actual Order Finish Date'. Using the obviously-named one yields garbage.
    """
    import pandas as pd

    frame = _read_excel(iw39_bytes, ['MaintActivityType', 'Main work center', 'System status',
                                     'Maintenance Plan', 'Actual Order Finish Date'])
    mes = frame[frame['Main work center'].fillna('').str.upper().str.startswith(MES_PREFIX)]
    finished = mes[~mes['System status'].map(is_plannable_status)]

    dates = pd.to_datetime(finished['Actual Order Finish Date'], errors='coerce')
    index = {}
    for plan, finished_on in zip(finished['Maintenance Plan'], dates):
        if plan is None or isinstance(plan, float) or pd.isna(finished_on):
            continue
        key = str(plan).strip()
        if not key:
            continue
        current = index.get(key)
        if current is None or finished_on > current:
            index[key] = finished_on
    return index


# ---------------------------------------------------------------------------
# Hourly PM priority
# ---------------------------------------------------------------------------
#
# Ali: "the interval is always 250 hrs between 2 maintenance [orders] for the
# same eqt and same maintenance plan number — the difference is in the additional
# task." So the 250HR/500HR/1000HR/2000HR variants are nested packages of ONE
# plan, and the service interval is 250 regardless of which package is named.
#
# That means the last completion of the plan is the right baseline, whichever
# package it was. Matching on the package level instead was tried and produced
# nonsense (RS115 showed 43 hours run against a 250-hour service).
#
# "Hours run" comes from IK17 counter readings: the reading nearest the last
# completion, subtracted from the latest reading.

PM_SERVICE_INTERVAL_HOURS = 250.0

# Ali's thresholds, 2026-08-22, expressed as hours STILL TO RUN before the
# service is due (negative = already past it):
#
#   past by more than 20h   -> top urgent
#   past by up to 20h       -> urgent
#   less than 30h to go     -> medium
#   30h or more to go       -> normal
#
# NOTE: the database CHECK constraint allows low/normal/high/urgent only — there
# is no "top urgent" and no "medium". So the top two tiers both return 'urgent'
# and Ali's "medium" is the existing 'high'. The distinction between the top two
# is NOT lost: `hours_past_due` is returned alongside, so the generator can rank
# within the urgent tier rather than treating a job 5 hours over the same as one
# 200 hours over.
HOURLY_TOP_URGENT_PAST_HOURS = 20.0
HOURLY_HIGH_WINDOW = 30.0


def hourly_pm_priority(hours_since_service,
                       interval=PM_SERVICE_INTERVAL_HOURS,
                       high_window=HOURLY_HIGH_WINDOW,
                       top_urgent_past=HOURLY_TOP_URGENT_PAST_HOURS):
    """Priority and severity for a running-hours PM.

    Returns (priority, hours_past_due) where hours_past_due is positive once the
    interval is exceeded. Returns (None, None) when it cannot be computed — a
    meter replacement makes the subtraction meaningless (TT034 reads 21,877 then
    3,920), and inventing a priority from that would be worse than admitting
    ignorance. Callers report those rather than hiding them.
    """
    if hours_since_service is None or hours_since_service < 0:
        return None, None

    hours_past_due = hours_since_service - interval

    if hours_past_due > top_urgent_past:
        return 'urgent', hours_past_due          # Ali's "top urgent"
    if hours_past_due >= 0:
        return 'urgent', hours_past_due
    if -hours_past_due < high_window:
        return 'high', hours_past_due            # Ali's "medium"
    return 'normal', hours_past_due


def build_meter_index(ik17_bytes):
    """Counter readings per machine, oldest first.

    Keyed on the plant code parsed out of Functional Location, same as orders,
    so the two sides join. Readings are sparse and occasionally non-monotonic
    (RS115 reads 30,271 on 7 May then 30,144 on 13 May), so callers should treat
    a negative difference as unknown rather than as a negative interval.
    """
    import pandas as pd

    frame = _read_excel(ik17_bytes, ['Date', 'Counter reading', 'Functional Location'])
    frame['equipment'] = frame['Functional Location'].map(extract_plant_code)
    frame['taken_on'] = pd.to_datetime(frame['Date'], errors='coerce')
    frame['reading'] = pd.to_numeric(
        frame['Counter reading'].astype(str).str.replace(',', '', regex=False), errors='coerce')
    frame = frame.dropna(subset=['equipment', 'taken_on', 'reading'])

    index = {}
    for equipment, group in frame.sort_values('taken_on').groupby('equipment'):
        index[equipment] = list(zip(group['taken_on'], group['reading']))
    return index


def hours_run_since(meter_index, equipment, since_date):
    """Counter hours run on `equipment` since `since_date`.

    Sums only the FORWARD movements between consecutive readings, ignoring any
    backward jump. A meter replacement is not missing data — it is data with a
    step in it, and subtracting first-from-last would throw away a whole machine.

    Real cases: TT033 read 8,974 then 3,161 on 12 May 2026, and TT034 read 21,912
    then 3,846 the SAME day — two tractors re-metered together. Naive subtraction
    gave -5,684 and -17,957; step-summing gives 119.6 and 87.7, which are usable.

    Known undercount: hours run between the last reading before a swap and the
    swap itself are unrecorded and therefore lost. Small next to losing the
    machine entirely, and it errs toward "not yet due" rather than raising a false
    alarm.

    Also absorbs the smaller non-monotonic glitches in this data — RS115 reads
    30,271 on 7 May then 30,144 on 13 May — which would otherwise poison a
    straight subtraction.

    Returns None when there is no reading at or before `since_date`, i.e. when
    there is genuinely nothing to measure from.
    """
    readings = meter_index.get(equipment)
    if not readings or since_date is None:
        return None
    if not any(taken_on <= since_date for taken_on, _ in readings):
        return None

    total = 0.0
    previous = None
    for taken_on, reading in readings:
        if taken_on < since_date:
            previous = reading
            continue
        if previous is not None:
            step = reading - previous
            if step > 0:
                total += step
        previous = reading
    return total


# ---------------------------------------------------------------------------
# Corrective priority (COM / DAM / ACD)
# ---------------------------------------------------------------------------
#
# Six layers, first match wins, then one promote-only adjustment.
#
# Everything else in SAP was tested and rejected as a signal:
#   * Order Priority      — 51% of open orders are "1", vs 2.3% of finished ones
#   * Notification Priority — 105 of 127 are "1-Extreme"
#   * Breakdown / Effect / Cond.aft.malfunctn — entirely empty in this SAP
#   * Required End        — median window from notification is 0 days, so it is
#                           just the creation date wearing a different hat
#   * Age since created   — median 296 days, 118 of 130 over 90. Nothing to sort
#                           with, so it became the TIEBREAKER instead (see below)
#
# The inspection signals (layers 1, 2, 5) come from the app's own database and
# are deliberately ranked ABOVE SAP's: an inspector stood in front of the machine
# last week, while SAP's severity fields are empty or diluted to meaninglessness.

CORRECTIVE_ACTIVITY_TYPES = ('COM', 'DAM', 'ACD')

BREAKDOWN_WINDOW_DAYS = 30
BREAKDOWN_URGENT_COUNT = 3
CLUSTER_HIGH_COUNT = 4

# Deliberately the MIDDLE width. Tested against the real 130 descriptions:
#   broad  (+cylinder, leak, engine, transmission, hydr) lifted 73/130 — 56%,
#          which is not a promotion, it is a new baseline
#   tight  (control & stopping only) lifted 40/130 but dropped
#          "Air System not working" and "Crack in Spreader T-Beam" — on a reach
#          stacker the spreader is what holds the container
#   middle lifts 51/130 and every lift is defensible
#
# Ali owns this list. It is configuration, and pruning a word is a one-line edit.
SAFETY_KEYWORDS = (
    'brake', 'steering', 'axle', 'suspension', 'wheel', 'tire', 'tires',
    'air', 'spreader',
)

_PROMOTE = {'normal': 'high', 'high': 'urgent', 'urgent': 'urgent'}


def mentions_safety_part(description):
    """Whether a job description names a part that affects control or stopping."""
    if not description:
        return False
    words = set(re.findall(r'[a-z]+', str(description).lower()))
    return bool(words & set(SAFETY_KEYWORDS))


def corrective_priority(*, equipment_assessment=None, defect_severity=None,
                        inspection_urgency=None, is_released=False,
                        breakdowns_30d=0, open_defects_on_equipment=0,
                        description=None):
    """Priority for a corrective job. Returns (priority, reason).

    `reason` is returned so the planner can show WHY a job is urgent. A ranking
    nobody can interrogate is a ranking nobody trusts — and with six layers, "it
    just is" would not survive its first argument in the yard.
    """
    priority = 'normal'
    reason = 'no signal'

    if equipment_assessment == 'stop':
        priority, reason = 'urgent', 'equipment assessed STOP'
    elif defect_severity == 'critical' or inspection_urgency == 3:
        priority, reason = 'urgent', 'critical inspection finding'
    elif is_released:
        priority, reason = 'urgent', 'released for work'
    elif breakdowns_30d >= BREAKDOWN_URGENT_COUNT:
        priority, reason = 'urgent', f'{breakdowns_30d} breakdowns in {BREAKDOWN_WINDOW_DAYS} days'
    elif equipment_assessment == 'monitor' or defect_severity == 'high':
        priority, reason = 'high', 'equipment under monitor'
    elif open_defects_on_equipment >= CLUSTER_HIGH_COUNT:
        priority, reason = 'high', f'{open_defects_on_equipment} open faults on this machine'

    # Promote-only, never demote. Missing a safety word leaves a job where the
    # other signals put it — harmless. Demoting one would bury a real brake fault
    # under a broken door handle, which is the failure that hurts somebody.
    if mentions_safety_part(description):
        promoted = _PROMOTE[priority]
        if promoted != priority:
            reason = 'safety part' if reason == 'no signal' else f'{reason} + safety part'
            priority = promoted

    return priority, reason


def build_breakdown_index(iw39_bytes, today=None, window_days=BREAKDOWN_WINDOW_DAYS):
    """Count of BDM orders per machine in the recent window.

    BDM is excluded from planning — breakdowns are handled reactively, not
    scheduled. But the COUNT is Ali's signal: a machine that keeps failing needs
    its outstanding defects looked at, because they may be why it keeps failing.

    Chosen at 30 days over 7: at a week the maximum any machine reaches is 2, so
    there is nothing to rank. At 30, RS118 (6) and RS113 (5) separate clearly.
    Real case for the signal: RS113 has ONE open defect — the cluster rule ranks
    it near the bottom — but broke down five times in a month.
    """
    import pandas as pd

    frame = _read_excel(iw39_bytes, ['MaintActivityType', 'Main work center',
                                     'Functional Location', 'Created on'])
    mes = frame[frame['Main work center'].fillna('').str.upper().str.startswith(MES_PREFIX)]
    breakdowns = mes[mes['MaintActivityType'] == 'BDM'].copy()

    now = _today_naive(today)
    created = pd.to_datetime(breakdowns['Created on'], errors='coerce')
    recent = breakdowns[created >= now - pd.Timedelta(days=window_days)]

    counts = {}
    for location in recent['Functional Location']:
        code = extract_plant_code(location)
        if code:
            counts[code] = counts.get(code, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Learned job durations
# ---------------------------------------------------------------------------
#
# IW39 carries no hours at all, and IW49 has them for 5,539 of 5,548 FINISHED
# MES orders but ZERO of the open ones — so the pool cannot be estimated from
# SAP directly. The history can estimate it instead.
#
# A flat default is not a neutral choice. Measured on Ali's real data, the true
# medians are PRM on TR = 18.0h, ECH = 14.6h, RS = 10.6h, TT = 8.2h against the
# app's blanket 4.0h. Planning a TR service as 4 hours puts four of them on one
# day and none of them finish.
#
# Median, not mean: one 40-hour disaster must not drag a whole family's estimate.

MIN_SAMPLES_FOR_DURATION = 5


def build_duration_index(iw39_bytes, hours_by_order):
    """Median hours per (activity type, equipment family), learned from history.

    Falls back through progressively wider buckets so there is always an answer
    and none of it is invented:
        (PRM, TT) -> (PRM, *) -> overall median -> DEFAULT_HOURS
    """
    import statistics

    frame = _read_excel(iw39_bytes, ['Order', 'MaintActivityType', 'Main work center',
                                     'System status', 'Functional Location'])
    mes = frame[frame['Main work center'].fillna('').str.upper().str.startswith(MES_PREFIX)]
    finished = mes[~mes['System status'].map(is_plannable_status)]

    by_pair, by_activity, overall = {}, {}, []
    for order, activity, location in zip(finished['Order'],
                                         finished['MaintActivityType'],
                                         finished['Functional Location']):
        if order is None or isinstance(order, float):
            continue
        hours = hours_by_order.get(str(order).strip())
        if not hours or hours <= 0:
            continue
        code = extract_plant_code(location)
        family = re.match(r'^([A-Z]{2,4})', code).group(1) if code else None
        act = (activity or '').strip().upper()
        overall.append(hours)
        by_activity.setdefault(act, []).append(hours)
        if family:
            by_pair.setdefault((act, family), []).append(hours)

    def median_of(samples):
        return round(statistics.median(samples), 2) if len(samples) >= MIN_SAMPLES_FOR_DURATION else None

    return {
        'by_pair': {k: median_of(v) for k, v in by_pair.items() if median_of(v)},
        'by_activity': {k: median_of(v) for k, v in by_activity.items() if median_of(v)},
        'overall': median_of(overall),
    }


def estimate_hours(duration_index, activity_type, plant_code):
    """Best learned estimate for one job, widening the bucket until something fits."""
    if not duration_index:
        return DEFAULT_HOURS
    act = (activity_type or '').strip().upper()
    family = None
    if plant_code:
        match = re.match(r'^([A-Z]{2,4})', plant_code)
        family = match.group(1) if match else None

    if family:
        specific = duration_index.get('by_pair', {}).get((act, family))
        if specific:
            return specific
    by_activity = duration_index.get('by_activity', {}).get(act)
    if by_activity:
        return by_activity
    return duration_index.get('overall') or DEFAULT_HOURS
