"""
Work Plan Service - Centralized business logic for work planning.
Handles templates, dependencies, job splitting, capacity management,
skill management, equipment restrictions, versioning, checklists,
conflict detection, validation, notifications, and reporting.
"""

import logging
import json
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any
from io import BytesIO

from sqlalchemy import func, and_, or_
from sqlalchemy.exc import IntegrityError

from app.extensions import db
from app.models import (
    JobTemplate, JobTemplateMaterial, JobTemplateChecklist,
    JobDependency, CapacityConfig, WorkerSkill, EquipmentRestriction,
    WorkPlanVersion, JobChecklistResponse, SchedulingConflict,
    WorkPlan, WorkPlanDay, WorkPlanJob, WorkPlanAssignment, WorkPlanMaterial,
    User, Equipment, Material, Leave
)
from app.exceptions.api_exceptions import (
    NotFoundError, ValidationError, ConflictError, BusinessError
)
from app.services.notification_service import NotificationService

logger = logging.getLogger(__name__)


class WorkPlanService:
    """Work planning business logic service."""

    # =========================================================================
    # TEMPLATE MANAGEMENT
    # =========================================================================

    @staticmethod
    def create_template(data: dict, created_by_id: int) -> JobTemplate:
        """Create a new job template.

        Args:
            data: Template data including name, job_type, estimated_hours, etc.
            created_by_id: ID of the user creating the template

        Returns:
            Created JobTemplate instance
        """
        # Validate required fields
        if not data.get('name'):
            raise ValidationError("Template name is required", field='name')
        if not data.get('job_type'):
            raise ValidationError("Job type is required", field='job_type')
        if data.get('job_type') not in ('pm', 'defect', 'inspection'):
            raise ValidationError("Invalid job type. Must be pm, defect, or inspection", field='job_type')
        if not data.get('estimated_hours') or data.get('estimated_hours') <= 0:
            raise ValidationError("Estimated hours must be positive", field='estimated_hours')

        template = JobTemplate(
            name=data.get('name'),
            name_ar=data.get('name_ar'),
            job_type=data.get('job_type'),
            equipment_id=data.get('equipment_id'),
            equipment_type=data.get('equipment_type'),
            berth=data.get('berth'),
            estimated_hours=data.get('estimated_hours'),
            priority=data.get('priority', 'normal'),
            description=data.get('description'),
            description_ar=data.get('description_ar'),
            recurrence_type=data.get('recurrence_type'),
            recurrence_day=data.get('recurrence_day'),
            default_team_size=data.get('default_team_size', 1),
            required_certifications=data.get('required_certifications'),
            is_active=True,
            created_by_id=created_by_id
        )

        db.session.add(template)
        db.session.commit()

        logger.info(f"Created job template: id={template.id} name={template.name}")
        return template

    @staticmethod
    def update_template(template_id: int, data: dict) -> JobTemplate:
        """Update a job template.

        Args:
            template_id: ID of the template to update
            data: Fields to update

        Returns:
            Updated JobTemplate instance
        """
        template = db.session.get(JobTemplate, template_id)
        if not template:
            raise NotFoundError(f"Job template with ID {template_id} not found")

        # Update allowed fields
        allowed_fields = [
            'name', 'name_ar', 'job_type', 'equipment_id', 'equipment_type',
            'berth', 'estimated_hours', 'priority', 'description', 'description_ar',
            'recurrence_type', 'recurrence_day', 'default_team_size',
            'required_certifications', 'is_active'
        ]

        for field in allowed_fields:
            if field in data:
                setattr(template, field, data[field])

        db.session.commit()

        logger.info(f"Updated job template: id={template_id}")
        return template

    @staticmethod
    def clone_template(template_id: int, new_name: str, new_name_ar: str = None) -> JobTemplate:
        """Clone a template including materials and checklist.

        Args:
            template_id: ID of the template to clone
            new_name: Name for the cloned template
            new_name_ar: Arabic name for the cloned template (optional)

        Returns:
            New cloned JobTemplate instance
        """
        original = db.session.get(JobTemplate, template_id)
        if not original:
            raise NotFoundError(f"Job template with ID {template_id} not found")

        # Create new template
        clone = JobTemplate(
            name=new_name,
            name_ar=new_name_ar,
            job_type=original.job_type,
            equipment_id=original.equipment_id,
            equipment_type=original.equipment_type,
            berth=original.berth,
            estimated_hours=original.estimated_hours,
            priority=original.priority,
            description=original.description,
            description_ar=original.description_ar,
            recurrence_type=original.recurrence_type,
            recurrence_day=original.recurrence_day,
            default_team_size=original.default_team_size,
            required_certifications=original.required_certifications,
            is_active=True,
            created_by_id=original.created_by_id
        )
        db.session.add(clone)
        db.session.flush()  # Get the clone ID

        # Clone materials
        for material in original.materials:
            clone_material = JobTemplateMaterial(
                template_id=clone.id,
                material_id=material.material_id,
                quantity=material.quantity,
                is_optional=material.is_optional
            )
            db.session.add(clone_material)

        # Clone checklist items
        for checklist in original.checklist_items:
            clone_checklist = JobTemplateChecklist(
                template_id=clone.id,
                item_code=checklist.item_code,
                question=checklist.question,
                question_ar=checklist.question_ar,
                answer_type=checklist.answer_type,
                is_required=checklist.is_required,
                order_index=checklist.order_index,
                fail_action=checklist.fail_action,
                fail_action_ar=checklist.fail_action_ar
            )
            db.session.add(clone_checklist)

        db.session.commit()

        logger.info(f"Cloned template {template_id} to new template {clone.id}")
        return clone

    @staticmethod
    def add_template_material(template_id: int, material_id: int, quantity: float,
                              is_optional: bool = False) -> JobTemplateMaterial:
        """Add material to template.

        Args:
            template_id: ID of the template
            material_id: ID of the material to add
            quantity: Required quantity
            is_optional: Whether the material is optional

        Returns:
            Created JobTemplateMaterial instance
        """
        template = db.session.get(JobTemplate, template_id)
        if not template:
            raise NotFoundError(f"Job template with ID {template_id} not found")

        material = db.session.get(Material, material_id)
        if not material:
            raise NotFoundError(f"Material with ID {material_id} not found")

        # Check if material already exists
        existing = JobTemplateMaterial.query.filter_by(
            template_id=template_id,
            material_id=material_id
        ).first()
        if existing:
            raise ConflictError(f"Material {material_id} already exists in template {template_id}")

        template_material = JobTemplateMaterial(
            template_id=template_id,
            material_id=material_id,
            quantity=quantity,
            is_optional=is_optional
        )
        db.session.add(template_material)
        db.session.commit()

        return template_material

    @staticmethod
    def add_template_checklist_item(template_id: int, data: dict) -> JobTemplateChecklist:
        """Add checklist item to template.

        Args:
            template_id: ID of the template
            data: Checklist item data

        Returns:
            Created JobTemplateChecklist instance
        """
        template = db.session.get(JobTemplate, template_id)
        if not template:
            raise NotFoundError(f"Job template with ID {template_id} not found")

        if not data.get('question'):
            raise ValidationError("Question is required", field='question')

        # Get next order index
        max_order = db.session.query(func.max(JobTemplateChecklist.order_index)).filter_by(
            template_id=template_id
        ).scalar() or 0

        checklist_item = JobTemplateChecklist(
            template_id=template_id,
            item_code=data.get('item_code'),
            question=data.get('question'),
            question_ar=data.get('question_ar'),
            answer_type=data.get('answer_type', 'pass_fail'),
            is_required=data.get('is_required', True),
            order_index=data.get('order_index', max_order + 1),
            fail_action=data.get('fail_action'),
            fail_action_ar=data.get('fail_action_ar')
        )
        db.session.add(checklist_item)
        db.session.commit()

        return checklist_item

    @staticmethod
    def get_matching_template(job_type: str, equipment_id: int = None,
                              equipment_type: str = None) -> Optional[JobTemplate]:
        """Find matching template for a job.

        Args:
            job_type: Type of job (pm, defect, inspection)
            equipment_id: Specific equipment ID (optional)
            equipment_type: Generic equipment type (optional)

        Returns:
            Matching JobTemplate or None
        """
        query = JobTemplate.query.filter_by(
            job_type=job_type,
            is_active=True
        )

        # Try to match specific equipment first
        if equipment_id:
            specific = query.filter_by(equipment_id=equipment_id).first()
            if specific:
                return specific

        # Then try equipment type
        if equipment_type:
            type_match = query.filter_by(equipment_type=equipment_type).first()
            if type_match:
                return type_match

        # Fall back to generic template (no equipment specified)
        return query.filter(
            JobTemplate.equipment_id.is_(None),
            JobTemplate.equipment_type.is_(None)
        ).first()

    # =========================================================================
    # DEPENDENCY MANAGEMENT
    # =========================================================================

    @staticmethod
    def add_dependency(job_id: int, depends_on_job_id: int,
                       dependency_type: str = 'finish_to_start',
                       lag_minutes: int = 0) -> JobDependency:
        """Add dependency between jobs.

        Args:
            job_id: ID of the job that has the dependency
            depends_on_job_id: ID of the job that must be completed/started first
            dependency_type: Type of dependency (finish_to_start, start_to_start)
            lag_minutes: Delay in minutes after the dependency is satisfied

        Returns:
            Created JobDependency instance

        Raises:
            ValidationError: If adding dependency would create a cycle
        """
        # Validate jobs exist
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        depends_on = db.session.get(WorkPlanJob, depends_on_job_id)
        if not depends_on:
            raise NotFoundError(f"Job with ID {depends_on_job_id} not found")

        # Check for self-dependency
        if job_id == depends_on_job_id:
            raise ValidationError("A job cannot depend on itself")

        # Check for circular dependency
        if WorkPlanService.check_circular_dependency(job_id, depends_on_job_id):
            raise ValidationError("Adding this dependency would create a circular dependency")

        # Validate dependency type
        if dependency_type not in ('finish_to_start', 'start_to_start'):
            raise ValidationError("Invalid dependency type. Must be finish_to_start or start_to_start")

        # Check if dependency already exists
        existing = JobDependency.query.filter_by(
            job_id=job_id,
            depends_on_job_id=depends_on_job_id
        ).first()
        if existing:
            raise ConflictError("This dependency already exists")

        dependency = JobDependency(
            job_id=job_id,
            depends_on_job_id=depends_on_job_id,
            dependency_type=dependency_type,
            lag_minutes=lag_minutes
        )
        db.session.add(dependency)
        db.session.commit()

        logger.info(f"Added dependency: job {job_id} depends on {depends_on_job_id}")
        return dependency

    @staticmethod
    def remove_dependency(dependency_id: int) -> bool:
        """Remove a dependency.

        Args:
            dependency_id: ID of the dependency to remove

        Returns:
            True if removed successfully
        """
        dependency = db.session.get(JobDependency, dependency_id)
        if not dependency:
            raise NotFoundError(f"Dependency with ID {dependency_id} not found")

        db.session.delete(dependency)
        db.session.commit()

        logger.info(f"Removed dependency: id={dependency_id}")
        return True

    @staticmethod
    def get_job_dependencies(job_id: int) -> dict:
        """Get job's dependencies (both depends_on and blocks).

        Args:
            job_id: ID of the job

        Returns:
            Dictionary with depends_on and blocks lists
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        # Jobs this job depends on
        depends_on = JobDependency.query.filter_by(job_id=job_id).all()

        # Jobs that depend on this job (this job blocks them)
        blocks = JobDependency.query.filter_by(depends_on_job_id=job_id).all()

        return {
            'depends_on': [d.to_dict() for d in depends_on],
            'blocks': [b.to_dict() for b in blocks]
        }

    @staticmethod
    def check_circular_dependency(job_id: int, depends_on_job_id: int) -> bool:
        """Check if adding dependency would create a cycle.

        Uses depth-first search to detect cycles.

        Args:
            job_id: ID of the job that would have the dependency
            depends_on_job_id: ID of the job it would depend on

        Returns:
            True if adding the dependency would create a cycle
        """
        # DFS to find if depends_on_job_id can reach job_id
        visited = set()
        stack = [depends_on_job_id]

        while stack:
            current = stack.pop()
            if current == job_id:
                return True  # Cycle detected

            if current in visited:
                continue
            visited.add(current)

            # Get all jobs that current depends on
            dependencies = JobDependency.query.filter_by(job_id=current).all()
            for dep in dependencies:
                stack.append(dep.depends_on_job_id)

        return False

    @staticmethod
    def get_dependency_chain(job_id: int) -> list:
        """Get full dependency chain for a job.

        Args:
            job_id: ID of the job

        Returns:
            Ordered list of jobs in the dependency chain
        """
        chain = []
        visited = set()

        def traverse(current_id: int, depth: int = 0):
            if current_id in visited:
                return
            visited.add(current_id)

            dependencies = JobDependency.query.filter_by(job_id=current_id).all()
            for dep in dependencies:
                traverse(dep.depends_on_job_id, depth + 1)

            job = db.session.get(WorkPlanJob, current_id)
            if job:
                chain.append({
                    'job_id': current_id,
                    'depth': depth,
                    'job': job.to_dict(compact=True)
                })

        traverse(job_id)
        return chain

    # =========================================================================
    # JOB SPLITTING
    # =========================================================================

    @staticmethod
    def split_job(job_id: int, parts: list) -> list:
        """Split a job across multiple days.

        Args:
            job_id: ID of the job to split
            parts: List of dicts with {day_id, hours} for each part

        Returns:
            List of new job parts
        """
        original_job = db.session.get(WorkPlanJob, job_id)
        if not original_job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        if original_job.is_split:
            raise ValidationError("Job is already split")

        if len(parts) < 2:
            raise ValidationError("At least 2 parts are required to split a job")

        # Validate total hours
        total_hours = sum(p.get('hours', 0) for p in parts)
        if abs(total_hours - original_job.estimated_hours) > 0.01:
            raise ValidationError(
                f"Split hours ({total_hours}) must equal original estimated hours ({original_job.estimated_hours})"
            )

        # Mark original as split
        original_job.is_split = True

        created_parts = []
        total_materials = len(original_job.materials)

        for i, part in enumerate(parts, start=1):
            day_id = part.get('day_id')
            hours = part.get('hours')

            # Validate day exists
            day = db.session.get(WorkPlanDay, day_id)
            if not day:
                raise NotFoundError(f"Work plan day with ID {day_id} not found")

            # Create new job part
            new_job = WorkPlanJob(
                work_plan_day_id=day_id,
                job_type=original_job.job_type,
                berth=original_job.berth,
                equipment_id=original_job.equipment_id,
                defect_id=original_job.defect_id,
                inspection_assignment_id=original_job.inspection_assignment_id,
                sap_order_number=original_job.sap_order_number,
                sap_order_type=original_job.sap_order_type,
                description=original_job.description,
                cycle_id=original_job.cycle_id,
                pm_template_id=original_job.pm_template_id,
                template_id=original_job.template_id,
                estimated_hours=hours,
                priority=original_job.priority,
                checklist_required=original_job.checklist_required,
                completion_photo_required=original_job.completion_photo_required,
                weather_sensitive=original_job.weather_sensitive,
                split_from_id=job_id,
                split_part=i,
                notes=original_job.notes
            )
            db.session.add(new_job)
            db.session.flush()

            # Distribute materials proportionally
            ratio = hours / original_job.estimated_hours
            for orig_material in original_job.materials:
                new_material = WorkPlanMaterial(
                    work_plan_job_id=new_job.id,
                    material_id=orig_material.material_id,
                    quantity=round(orig_material.quantity * ratio, 2),
                    from_kit_id=orig_material.from_kit_id
                )
                db.session.add(new_material)

            # Copy assignments to all parts
            for orig_assignment in original_job.assignments:
                new_assignment = WorkPlanAssignment(
                    work_plan_job_id=new_job.id,
                    user_id=orig_assignment.user_id,
                    is_lead=orig_assignment.is_lead
                )
                db.session.add(new_assignment)

            created_parts.append(new_job)

        db.session.commit()

        logger.info(f"Split job {job_id} into {len(parts)} parts")
        return created_parts

    @staticmethod
    def merge_split_jobs(original_job_id: int) -> WorkPlanJob:
        """Merge split job parts back into one.

        Args:
            original_job_id: ID of the original job that was split

        Returns:
            The merged WorkPlanJob
        """
        original = db.session.get(WorkPlanJob, original_job_id)
        if not original:
            raise NotFoundError(f"Job with ID {original_job_id} not found")

        if not original.is_split:
            raise ValidationError("Job is not split")

        # Find all split parts
        split_parts = WorkPlanJob.query.filter_by(split_from_id=original_job_id).all()
        if not split_parts:
            raise ValidationError("No split parts found for this job")

        # Aggregate data from parts
        total_hours = sum(p.estimated_hours for p in split_parts)

        # Delete split parts and their materials/assignments
        for part in split_parts:
            # Delete materials
            WorkPlanMaterial.query.filter_by(work_plan_job_id=part.id).delete()
            # Delete assignments
            WorkPlanAssignment.query.filter_by(work_plan_job_id=part.id).delete()
            # Delete the part
            db.session.delete(part)

        # Update original job
        original.is_split = False
        original.estimated_hours = total_hours

        db.session.commit()

        logger.info(f"Merged split jobs back to original job {original_job_id}")
        return original

    # =========================================================================
    # CAPACITY MANAGEMENT
    # =========================================================================

    @staticmethod
    def get_capacity_config(role: str = None, shift: str = None) -> Optional[CapacityConfig]:
        """Get applicable capacity config.

        Args:
            role: User role filter
            shift: Shift filter (day/night)

        Returns:
            Most specific matching CapacityConfig or None
        """
        query = CapacityConfig.query.filter_by(is_active=True)

        # Try to find most specific match
        if role and shift:
            config = query.filter_by(role=role, shift=shift).first()
            if config:
                return config

        # Try role-specific
        if role:
            config = query.filter_by(role=role, shift=None).first()
            if config:
                return config

        # Try shift-specific
        if shift:
            config = query.filter_by(role=None, shift=shift).first()
            if config:
                return config

        # Fall back to default (no role or shift specified)
        return query.filter_by(role=None, shift=None).first()

    @staticmethod
    def check_capacity_violation(user_id: int, day_id: int,
                                 additional_hours: float = 0) -> dict:
        """Check if user exceeds capacity for a day.

        Args:
            user_id: ID of the user
            day_id: ID of the work plan day
            additional_hours: Hours to add (for checking before assignment)

        Returns:
            Dictionary with violation status and details
        """
        user = db.session.get(User, user_id)
        if not user:
            raise NotFoundError(f"User with ID {user_id} not found")

        day = db.session.get(WorkPlanDay, day_id)
        if not day:
            raise NotFoundError(f"Work plan day with ID {day_id} not found")

        # Get user's current assignments for this day
        assignments = WorkPlanAssignment.query.join(WorkPlanJob).filter(
            WorkPlanAssignment.user_id == user_id,
            WorkPlanJob.work_plan_day_id == day_id
        ).all()

        current_hours = sum(a.job.estimated_hours for a in assignments)
        total_hours = current_hours + additional_hours

        # Get capacity config for user
        config = WorkPlanService.get_capacity_config(user.role, user.shift)
        max_hours = config.max_hours_per_day if config else 8
        overtime_threshold = config.overtime_threshold_hours if config else 8
        max_overtime = config.max_overtime_hours if config else 4

        violated = total_hours > (max_hours + max_overtime)
        overtime_hours = max(0, total_hours - overtime_threshold)

        return {
            'violated': violated,
            'current_hours': current_hours,
            'additional_hours': additional_hours,
            'total_hours': total_hours,
            'max_hours': max_hours,
            'overtime_threshold': overtime_threshold,
            'max_overtime': max_overtime,
            'overtime_hours': overtime_hours
        }

    @staticmethod
    def get_available_capacity(day_id: int, role: str = None,
                               shift: str = None) -> list:
        """Get available capacity per user for a day.

        Args:
            day_id: ID of the work plan day
            role: Filter by user role
            shift: Filter by user shift

        Returns:
            List of dicts with user_id, available_hours, current_jobs
        """
        day = db.session.get(WorkPlanDay, day_id)
        if not day:
            raise NotFoundError(f"Work plan day with ID {day_id} not found")

        # Build user query
        user_query = User.query.filter_by(is_active=True)
        if role:
            user_query = user_query.filter(
                or_(User.role == role, User.minor_role == role)
            )
        if shift:
            user_query = user_query.filter_by(shift=shift)

        users = user_query.all()
        result = []

        for user in users:
            # Skip users on leave
            if user.is_on_leave:
                continue

            # Check for approved leave on this day
            leave = Leave.query.filter(
                Leave.user_id == user.id,
                Leave.status == 'approved',
                Leave.date_from <= day.date,
                Leave.date_to >= day.date
            ).first()
            if leave:
                continue

            # Get current assignments
            assignments = WorkPlanAssignment.query.join(WorkPlanJob).filter(
                WorkPlanAssignment.user_id == user.id,
                WorkPlanJob.work_plan_day_id == day_id
            ).all()

            current_hours = sum(a.job.estimated_hours for a in assignments)
            current_jobs = len(assignments)

            # Get capacity config
            config = WorkPlanService.get_capacity_config(user.role, user.shift)
            max_hours = config.max_hours_per_day if config else 8
            max_jobs = config.max_jobs_per_day if config else 5

            available_hours = max_hours - current_hours
            available_jobs = max_jobs - current_jobs

            result.append({
                'user_id': user.id,
                'user_name': user.full_name,
                'role': user.role,
                'shift': user.shift,
                'available_hours': max(0, available_hours),
                'available_jobs': max(0, available_jobs),
                'current_hours': current_hours,
                'current_jobs': current_jobs,
                'max_hours': max_hours,
                'max_jobs': max_jobs
            })

        # Sort by available hours descending
        result.sort(key=lambda x: x['available_hours'], reverse=True)
        return result

    # =========================================================================
    # SKILL MANAGEMENT
    # =========================================================================

    @staticmethod
    def add_skill(user_id: int, data: dict, added_by_id: int) -> WorkerSkill:
        """Add skill/certification to user.

        Args:
            user_id: ID of the user
            data: Skill data
            added_by_id: ID of user adding the skill

        Returns:
            Created WorkerSkill instance
        """
        user = db.session.get(User, user_id)
        if not user:
            raise NotFoundError(f"User with ID {user_id} not found")

        if not data.get('skill_name'):
            raise ValidationError("Skill name is required", field='skill_name')

        # Check for duplicate
        existing = WorkerSkill.query.filter_by(
            user_id=user_id,
            skill_name=data.get('skill_name')
        ).first()
        if existing:
            raise ConflictError(f"User already has skill '{data.get('skill_name')}'")

        # Parse dates
        issued_date = None
        expiry_date = None
        if data.get('issued_date'):
            issued_date = datetime.strptime(data['issued_date'], '%Y-%m-%d').date()
        if data.get('expiry_date'):
            expiry_date = datetime.strptime(data['expiry_date'], '%Y-%m-%d').date()

        skill = WorkerSkill(
            user_id=user_id,
            skill_name=data.get('skill_name'),
            skill_level=data.get('skill_level', 1),
            certification_name=data.get('certification_name'),
            certification_number=data.get('certification_number'),
            issued_date=issued_date,
            expiry_date=expiry_date,
            issuing_authority=data.get('issuing_authority'),
            document_file_id=data.get('document_file_id'),
            is_verified=False
        )
        db.session.add(skill)
        db.session.commit()

        logger.info(f"Added skill '{data.get('skill_name')}' to user {user_id}")
        return skill

    @staticmethod
    def verify_skill(skill_id: int, verified_by_id: int) -> WorkerSkill:
        """Mark skill as verified.

        Args:
            skill_id: ID of the skill to verify
            verified_by_id: ID of user verifying the skill

        Returns:
            Updated WorkerSkill instance
        """
        skill = db.session.get(WorkerSkill, skill_id)
        if not skill:
            raise NotFoundError(f"Skill with ID {skill_id} not found")

        skill.is_verified = True
        skill.verified_by_id = verified_by_id
        skill.verified_at = datetime.utcnow()

        db.session.commit()

        logger.info(f"Verified skill {skill_id} by user {verified_by_id}")
        return skill

    @staticmethod
    def get_expiring_certifications(days_ahead: int = 30) -> list:
        """Get certifications expiring soon.

        Args:
            days_ahead: Number of days to look ahead

        Returns:
            List of WorkerSkill instances expiring within the period
        """
        cutoff_date = datetime.utcnow().date() + timedelta(days=days_ahead)

        skills = WorkerSkill.query.filter(
            WorkerSkill.expiry_date.isnot(None),
            WorkerSkill.expiry_date <= cutoff_date,
            WorkerSkill.expiry_date >= datetime.utcnow().date()
        ).all()

        return [s.to_dict() for s in skills]

    @staticmethod
    def check_skill_requirements(job_id: int, user_id: int) -> dict:
        """Check if user meets job's skill requirements.

        Args:
            job_id: ID of the job
            user_id: ID of the user

        Returns:
            Dictionary with meets_requirements and missing_skills
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        user = db.session.get(User, user_id)
        if not user:
            raise NotFoundError(f"User with ID {user_id} not found")

        # Get required certifications from template
        required_skills = []
        if job.template and job.template.required_certifications:
            required_skills = job.template.required_certifications

        if not required_skills:
            return {'meets_requirements': True, 'missing_skills': []}

        # Get user's verified, non-expired skills
        user_skills = WorkerSkill.query.filter(
            WorkerSkill.user_id == user_id,
            WorkerSkill.is_verified == True,
            or_(
                WorkerSkill.expiry_date.is_(None),
                WorkerSkill.expiry_date >= datetime.utcnow().date()
            )
        ).all()

        user_skill_names = {s.skill_name.lower() for s in user_skills}

        # Check for missing skills
        missing_skills = []
        for required in required_skills:
            if required.lower() not in user_skill_names:
                missing_skills.append(required)

        return {
            'meets_requirements': len(missing_skills) == 0,
            'missing_skills': missing_skills,
            'required_skills': required_skills,
            'user_skills': [s.skill_name for s in user_skills]
        }

    @staticmethod
    def get_qualified_workers(required_skills: list) -> list:
        """Get workers with required skills.

        Args:
            required_skills: List of skill names required

        Returns:
            List of qualified User instances
        """
        if not required_skills:
            return User.query.filter_by(is_active=True).all()

        # Get users who have all required skills (verified and not expired)
        qualified_users = []
        users = User.query.filter_by(is_active=True).all()

        for user in users:
            user_skills = WorkerSkill.query.filter(
                WorkerSkill.user_id == user.id,
                WorkerSkill.is_verified == True,
                or_(
                    WorkerSkill.expiry_date.is_(None),
                    WorkerSkill.expiry_date >= datetime.utcnow().date()
                )
            ).all()

            user_skill_names = {s.skill_name.lower() for s in user_skills}

            # Check if user has all required skills
            has_all = all(
                req.lower() in user_skill_names
                for req in required_skills
            )

            if has_all:
                qualified_users.append({
                    'user': user.to_dict(),
                    'skills': [s.to_dict() for s in user_skills]
                })

        return qualified_users

    # =========================================================================
    # EQUIPMENT RESTRICTIONS
    # =========================================================================

    @staticmethod
    def add_equipment_restriction(data: dict, created_by_id: int) -> EquipmentRestriction:
        """Add restriction to equipment.

        Args:
            data: Restriction data
            created_by_id: ID of user creating the restriction

        Returns:
            Created EquipmentRestriction instance
        """
        equipment_id = data.get('equipment_id')
        if not equipment_id:
            raise ValidationError("Equipment ID is required", field='equipment_id')

        equipment = db.session.get(Equipment, equipment_id)
        if not equipment:
            raise NotFoundError(f"Equipment with ID {equipment_id} not found")

        restriction_type = data.get('restriction_type')
        if restriction_type not in ('blackout', 'crew_size', 'skill_required', 'shift_only'):
            raise ValidationError(
                "Invalid restriction type. Must be blackout, crew_size, skill_required, or shift_only",
                field='restriction_type'
            )

        # Parse dates
        start_date = None
        end_date = None
        if data.get('start_date'):
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        if data.get('end_date'):
            end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

        restriction = EquipmentRestriction(
            equipment_id=equipment_id,
            restriction_type=restriction_type,
            value=data.get('value'),
            reason=data.get('reason'),
            start_date=start_date,
            end_date=end_date,
            is_permanent=data.get('is_permanent', False),
            is_active=True,
            created_by_id=created_by_id
        )
        db.session.add(restriction)
        db.session.commit()

        logger.info(f"Added restriction to equipment {equipment_id}: {restriction_type}")
        return restriction

    @staticmethod
    def check_equipment_restrictions(equipment_id: int, check_date: date,
                                     user_ids: list = None) -> dict:
        """Check equipment restrictions.

        Args:
            equipment_id: ID of the equipment
            check_date: Date to check restrictions for
            user_ids: List of user IDs to check (optional)

        Returns:
            Dictionary with allowed status and violations
        """
        equipment = db.session.get(Equipment, equipment_id)
        if not equipment:
            raise NotFoundError(f"Equipment with ID {equipment_id} not found")

        # Get active restrictions for this equipment on this date
        restrictions = EquipmentRestriction.query.filter(
            EquipmentRestriction.equipment_id == equipment_id,
            EquipmentRestriction.is_active == True,
            or_(
                EquipmentRestriction.is_permanent == True,
                and_(
                    or_(
                        EquipmentRestriction.start_date.is_(None),
                        EquipmentRestriction.start_date <= check_date
                    ),
                    or_(
                        EquipmentRestriction.end_date.is_(None),
                        EquipmentRestriction.end_date >= check_date
                    )
                )
            )
        ).all()

        violations = []
        allowed = True

        for restriction in restrictions:
            if restriction.restriction_type == 'blackout':
                violations.append({
                    'type': 'blackout',
                    'restriction_id': restriction.id,
                    'reason': restriction.reason or 'Equipment is not available',
                    'severity': 'error'
                })
                allowed = False

            elif restriction.restriction_type == 'crew_size' and user_ids:
                value = restriction.value or {}
                min_crew = value.get('min', 1)
                max_crew = value.get('max')

                if len(user_ids) < min_crew:
                    violations.append({
                        'type': 'crew_size',
                        'restriction_id': restriction.id,
                        'reason': f'Minimum crew size is {min_crew}, but only {len(user_ids)} assigned',
                        'severity': 'error'
                    })
                    allowed = False

                if max_crew and len(user_ids) > max_crew:
                    violations.append({
                        'type': 'crew_size',
                        'restriction_id': restriction.id,
                        'reason': f'Maximum crew size is {max_crew}, but {len(user_ids)} assigned',
                        'severity': 'warning'
                    })

            elif restriction.restriction_type == 'skill_required' and user_ids:
                value = restriction.value or {}
                required_skills = value.get('skills', [])

                for user_id in user_ids:
                    result = WorkPlanService.check_skill_requirements_for_user(user_id, required_skills)
                    if not result['meets_requirements']:
                        violations.append({
                            'type': 'skill_required',
                            'restriction_id': restriction.id,
                            'user_id': user_id,
                            'reason': f'User {user_id} missing skills: {", ".join(result["missing_skills"])}',
                            'severity': 'error'
                        })
                        allowed = False

            elif restriction.restriction_type == 'shift_only' and user_ids:
                value = restriction.value or {}
                required_shift = value.get('shift')

                for user_id in user_ids:
                    user = db.session.get(User, user_id)
                    if user and user.shift != required_shift:
                        violations.append({
                            'type': 'shift_only',
                            'restriction_id': restriction.id,
                            'user_id': user_id,
                            'reason': f'User {user_id} is on {user.shift} shift, but {required_shift} shift required',
                            'severity': 'warning'
                        })

        return {
            'allowed': allowed,
            'violations': violations
        }

    @staticmethod
    def check_skill_requirements_for_user(user_id: int, required_skills: list) -> dict:
        """Check if user has required skills.

        Helper method for equipment restriction checks.
        """
        user_skills = WorkerSkill.query.filter(
            WorkerSkill.user_id == user_id,
            WorkerSkill.is_verified == True,
            or_(
                WorkerSkill.expiry_date.is_(None),
                WorkerSkill.expiry_date >= datetime.utcnow().date()
            )
        ).all()

        user_skill_names = {s.skill_name.lower() for s in user_skills}

        missing_skills = []
        for required in required_skills:
            if required.lower() not in user_skill_names:
                missing_skills.append(required)

        return {
            'meets_requirements': len(missing_skills) == 0,
            'missing_skills': missing_skills
        }

    # =========================================================================
    # VERSION CONTROL
    # =========================================================================

    @staticmethod
    def create_version(plan_id: int, change_type: str, change_summary: str,
                       created_by_id: int) -> WorkPlanVersion:
        """Create a new version snapshot.

        Args:
            plan_id: ID of the work plan
            change_type: Type of change (created, jobs_added, jobs_moved, published, jobs_removed, updated)
            change_summary: Summary of changes
            created_by_id: ID of user creating the version

        Returns:
            Created WorkPlanVersion instance
        """
        plan = db.session.get(WorkPlan, plan_id)
        if not plan:
            raise NotFoundError(f"Work plan with ID {plan_id} not found")

        # Get next version number
        max_version = db.session.query(func.max(WorkPlanVersion.version_number)).filter_by(
            work_plan_id=plan_id
        ).scalar() or 0

        # Create snapshot of current plan state
        snapshot_data = WorkPlanService._create_plan_snapshot(plan)

        version = WorkPlanVersion(
            work_plan_id=plan_id,
            version_number=max_version + 1,
            snapshot_data=snapshot_data,
            change_type=change_type,
            change_summary=change_summary,
            created_by_id=created_by_id
        )
        db.session.add(version)
        db.session.commit()

        logger.info(f"Created version {version.version_number} for plan {plan_id}")
        return version

    @staticmethod
    def _create_plan_snapshot(plan: WorkPlan) -> dict:
        """Create a JSON snapshot of plan state."""
        return {
            'plan': {
                'id': plan.id,
                'week_start': plan.week_start.isoformat(),
                'week_end': plan.week_end.isoformat(),
                'status': plan.status,
                'notes': plan.notes
            },
            'days': [
                {
                    'id': day.id,
                    'date': day.date.isoformat(),
                    'notes': day.notes,
                    'jobs': [
                        {
                            'id': job.id,
                            'job_type': job.job_type,
                            'berth': job.berth,
                            'equipment_id': job.equipment_id,
                            'estimated_hours': job.estimated_hours,
                            'priority': job.priority,
                            'description': job.description,
                            'assignments': [
                                {'user_id': a.user_id, 'is_lead': a.is_lead}
                                for a in job.assignments
                            ],
                            'materials': [
                                {'material_id': m.material_id, 'quantity': m.quantity}
                                for m in job.materials
                            ]
                        }
                        for job in day.jobs
                    ]
                }
                for day in plan.days
            ],
            'snapshot_time': datetime.utcnow().isoformat()
        }

    @staticmethod
    def restore_version(plan_id: int, version_number: int,
                        restored_by_id: int) -> WorkPlan:
        """Restore plan to a previous version.

        Args:
            plan_id: ID of the work plan
            version_number: Version number to restore
            restored_by_id: ID of user restoring the version

        Returns:
            Updated WorkPlan instance
        """
        plan = db.session.get(WorkPlan, plan_id)
        if not plan:
            raise NotFoundError(f"Work plan with ID {plan_id} not found")

        version = WorkPlanVersion.query.filter_by(
            work_plan_id=plan_id,
            version_number=version_number
        ).first()
        if not version:
            raise NotFoundError(f"Version {version_number} not found for plan {plan_id}")

        # Create a new version before restoring (to preserve current state)
        WorkPlanService.create_version(
            plan_id, 'updated',
            f'Before restoring to version {version_number}',
            restored_by_id
        )

        snapshot = version.snapshot_data

        # Clear current jobs
        for day in plan.days:
            for job in day.jobs[:]:
                WorkPlanMaterial.query.filter_by(work_plan_job_id=job.id).delete()
                WorkPlanAssignment.query.filter_by(work_plan_job_id=job.id).delete()
                db.session.delete(job)

        # Restore from snapshot
        plan.notes = snapshot.get('plan', {}).get('notes')

        for day_data in snapshot.get('days', []):
            day = WorkPlanDay.query.filter_by(
                work_plan_id=plan_id,
                date=datetime.strptime(day_data['date'], '%Y-%m-%d').date()
            ).first()

            if day:
                day.notes = day_data.get('notes')

                for job_data in day_data.get('jobs', []):
                    job = WorkPlanJob(
                        work_plan_day_id=day.id,
                        job_type=job_data.get('job_type'),
                        berth=job_data.get('berth'),
                        equipment_id=job_data.get('equipment_id'),
                        estimated_hours=job_data.get('estimated_hours'),
                        priority=job_data.get('priority', 'normal'),
                        description=job_data.get('description')
                    )
                    db.session.add(job)
                    db.session.flush()

                    # Restore assignments
                    for assign_data in job_data.get('assignments', []):
                        assignment = WorkPlanAssignment(
                            work_plan_job_id=job.id,
                            user_id=assign_data.get('user_id'),
                            is_lead=assign_data.get('is_lead', False)
                        )
                        db.session.add(assignment)

                    # Restore materials
                    for mat_data in job_data.get('materials', []):
                        material = WorkPlanMaterial(
                            work_plan_job_id=job.id,
                            material_id=mat_data.get('material_id'),
                            quantity=mat_data.get('quantity')
                        )
                        db.session.add(material)

        # Create version for the restore
        WorkPlanService.create_version(
            plan_id, 'updated',
            f'Restored from version {version_number}',
            restored_by_id
        )

        db.session.commit()

        logger.info(f"Restored plan {plan_id} to version {version_number}")
        return plan

    @staticmethod
    def get_version_diff(plan_id: int, version_a: int, version_b: int) -> dict:
        """Compare two versions.

        Args:
            plan_id: ID of the work plan
            version_a: First version number
            version_b: Second version number

        Returns:
            Dictionary with added_jobs, removed_jobs, changed_jobs, assignment_changes
        """
        plan = db.session.get(WorkPlan, plan_id)
        if not plan:
            raise NotFoundError(f"Work plan with ID {plan_id} not found")

        ver_a = WorkPlanVersion.query.filter_by(
            work_plan_id=plan_id, version_number=version_a
        ).first()
        ver_b = WorkPlanVersion.query.filter_by(
            work_plan_id=plan_id, version_number=version_b
        ).first()

        if not ver_a:
            raise NotFoundError(f"Version {version_a} not found")
        if not ver_b:
            raise NotFoundError(f"Version {version_b} not found")

        # Extract jobs from both versions
        jobs_a = {}
        jobs_b = {}

        for day in ver_a.snapshot_data.get('days', []):
            for job in day.get('jobs', []):
                jobs_a[job['id']] = job

        for day in ver_b.snapshot_data.get('days', []):
            for job in day.get('jobs', []):
                jobs_b[job['id']] = job

        ids_a = set(jobs_a.keys())
        ids_b = set(jobs_b.keys())

        added_jobs = [jobs_b[id] for id in (ids_b - ids_a)]
        removed_jobs = [jobs_a[id] for id in (ids_a - ids_b)]

        # Check for changes in common jobs
        changed_jobs = []
        assignment_changes = []

        for job_id in (ids_a & ids_b):
            job_a = jobs_a[job_id]
            job_b = jobs_b[job_id]

            changes = {}
            for field in ['berth', 'equipment_id', 'estimated_hours', 'priority', 'description']:
                if job_a.get(field) != job_b.get(field):
                    changes[field] = {'from': job_a.get(field), 'to': job_b.get(field)}

            if changes:
                changed_jobs.append({
                    'job_id': job_id,
                    'changes': changes
                })

            # Check assignment changes
            assigns_a = {a['user_id'] for a in job_a.get('assignments', [])}
            assigns_b = {a['user_id'] for a in job_b.get('assignments', [])}

            if assigns_a != assigns_b:
                assignment_changes.append({
                    'job_id': job_id,
                    'added_users': list(assigns_b - assigns_a),
                    'removed_users': list(assigns_a - assigns_b)
                })

        return {
            'version_a': version_a,
            'version_b': version_b,
            'added_jobs': added_jobs,
            'removed_jobs': removed_jobs,
            'changed_jobs': changed_jobs,
            'assignment_changes': assignment_changes
        }

    # =========================================================================
    # CHECKLIST MANAGEMENT
    # =========================================================================

    @staticmethod
    def get_job_checklist(job_id: int) -> list:
        """Get checklist for a job (from template).

        Args:
            job_id: ID of the job

        Returns:
            List of checklist items with response status
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        if not job.template_id:
            return []

        # Get template checklist items
        items = JobTemplateChecklist.query.filter_by(
            template_id=job.template_id
        ).order_by(JobTemplateChecklist.order_index).all()

        # Get existing responses
        responses = {r.checklist_item_id: r for r in job.checklist_responses}

        result = []
        for item in items:
            response = responses.get(item.id)
            result.append({
                'item': item.to_dict(),
                'response': response.to_dict() if response else None,
                'answered': response is not None
            })

        return result

    @staticmethod
    def submit_checklist_response(job_id: int, item_id: int, data: dict,
                                  answered_by_id: int) -> JobChecklistResponse:
        """Submit a checklist response.

        Args:
            job_id: ID of the job
            item_id: ID of the checklist item
            data: Response data (answer_value, notes, photo_file_id)
            answered_by_id: ID of user submitting the response

        Returns:
            Created or updated JobChecklistResponse instance
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        item = db.session.get(JobTemplateChecklist, item_id)
        if not item:
            raise NotFoundError(f"Checklist item with ID {item_id} not found")

        answer_value = data.get('answer_value')
        if not answer_value and item.is_required:
            raise ValidationError("Answer is required for this item", field='answer_value')

        # Determine pass/fail status
        is_passed = None
        if item.answer_type == 'pass_fail':
            is_passed = answer_value.lower() == 'pass' if answer_value else None
        elif item.answer_type == 'yes_no':
            is_passed = answer_value.lower() == 'yes' if answer_value else None

        # Check for existing response
        existing = JobChecklistResponse.query.filter_by(
            work_plan_job_id=job_id,
            checklist_item_id=item_id
        ).first()

        if existing:
            # Update existing response
            existing.answer_value = answer_value
            existing.is_passed = is_passed
            existing.notes = data.get('notes')
            existing.photo_file_id = data.get('photo_file_id')
            existing.answered_by_id = answered_by_id
            existing.answered_at = datetime.utcnow()
            response = existing
        else:
            # Create new response
            response = JobChecklistResponse(
                work_plan_job_id=job_id,
                checklist_item_id=item_id,
                question=item.question,
                answer_type=item.answer_type,
                answer_value=answer_value,
                is_passed=is_passed,
                notes=data.get('notes'),
                photo_file_id=data.get('photo_file_id'),
                answered_by_id=answered_by_id
            )
            db.session.add(response)

        db.session.commit()

        logger.info(f"Submitted checklist response for job {job_id} item {item_id}")
        return response

    @staticmethod
    def validate_checklist_completion(job_id: int) -> dict:
        """Validate all required checklist items are answered.

        Args:
            job_id: ID of the job

        Returns:
            Dictionary with complete status, missing_items, and failed_items
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        if not job.template_id:
            return {'complete': True, 'missing_items': [], 'failed_items': []}

        # Get required checklist items
        required_items = JobTemplateChecklist.query.filter_by(
            template_id=job.template_id,
            is_required=True
        ).all()

        # Get responses
        responses = {r.checklist_item_id: r for r in job.checklist_responses}

        missing_items = []
        failed_items = []

        for item in required_items:
            response = responses.get(item.id)
            if not response:
                missing_items.append(item.to_dict())
            elif response.is_passed is False:
                failed_items.append({
                    'item': item.to_dict(),
                    'response': response.to_dict()
                })

        return {
            'complete': len(missing_items) == 0,
            'missing_items': missing_items,
            'failed_items': failed_items
        }

    @staticmethod
    def mark_checklist_complete(job_id: int) -> bool:
        """Mark job checklist as complete (validates first).

        Args:
            job_id: ID of the job

        Returns:
            True if marked complete

        Raises:
            ValidationError: If checklist is not complete
        """
        validation = WorkPlanService.validate_checklist_completion(job_id)

        if not validation['complete']:
            raise ValidationError(
                f"Checklist incomplete. Missing {len(validation['missing_items'])} required items."
            )

        job = db.session.get(WorkPlanJob, job_id)
        job.checklist_completed = True
        db.session.commit()

        logger.info(f"Marked checklist complete for job {job_id}")
        return True

    # =========================================================================
    # CONFLICT DETECTION
    # =========================================================================

    @staticmethod
    def detect_conflicts(plan_id: int) -> list:
        """Detect all scheduling conflicts for a plan.

        Checks: capacity, skill, equipment, dependency, overlap

        Args:
            plan_id: ID of the work plan

        Returns:
            List of detected conflicts
        """
        plan = db.session.get(WorkPlan, plan_id)
        if not plan:
            raise NotFoundError(f"Work plan with ID {plan_id} not found")

        conflicts = []

        # Clear existing unresolved conflicts for this plan
        SchedulingConflict.query.filter_by(
            work_plan_id=plan_id,
            resolved_at=None,
            is_ignored=False
        ).delete()

        for day in plan.days:
            # Track user assignments per day
            user_hours = {}
            user_jobs = {}

            for job in day.jobs:
                job_conflicts = WorkPlanService._detect_job_conflicts_internal(
                    job, day, user_hours, user_jobs
                )
                conflicts.extend(job_conflicts)

        # Save conflicts to database
        for conflict_data in conflicts:
            conflict = SchedulingConflict(
                work_plan_id=plan_id,
                conflict_type=conflict_data['type'],
                severity=conflict_data['severity'],
                description=conflict_data['description'],
                affected_job_ids=conflict_data.get('job_ids'),
                affected_user_ids=conflict_data.get('user_ids')
            )
            db.session.add(conflict)

        db.session.commit()

        logger.info(f"Detected {len(conflicts)} conflicts for plan {plan_id}")
        return conflicts

    @staticmethod
    def _detect_job_conflicts_internal(job: WorkPlanJob, day: WorkPlanDay,
                                       user_hours: dict, user_jobs: dict) -> list:
        """Internal method to detect conflicts for a single job."""
        conflicts = []

        for assignment in job.assignments:
            user_id = assignment.user_id
            user = assignment.user

            # Initialize tracking
            if user_id not in user_hours:
                user_hours[user_id] = 0
            if user_id not in user_jobs:
                user_jobs[user_id] = []

            user_hours[user_id] += job.estimated_hours
            user_jobs[user_id].append(job.id)

            # Capacity check
            capacity = WorkPlanService.check_capacity_violation(
                user_id, day.id, 0  # Already added hours
            )
            if capacity['violated']:
                conflicts.append({
                    'type': 'capacity',
                    'severity': 'warning',
                    'description': f'{user.full_name} exceeds capacity with {capacity["total_hours"]} hours (max {capacity["max_hours"]})',
                    'job_ids': [job.id],
                    'user_ids': [user_id]
                })

            # Skill check
            if job.template_id:
                skill_check = WorkPlanService.check_skill_requirements(job.id, user_id)
                if not skill_check['meets_requirements']:
                    conflicts.append({
                        'type': 'skill',
                        'severity': 'error',
                        'description': f'{user.full_name} missing required skills: {", ".join(skill_check["missing_skills"])}',
                        'job_ids': [job.id],
                        'user_ids': [user_id]
                    })

        # Equipment restriction check
        if job.equipment_id:
            user_ids = [a.user_id for a in job.assignments]
            restriction_check = WorkPlanService.check_equipment_restrictions(
                job.equipment_id, day.date, user_ids
            )
            if not restriction_check['allowed']:
                for violation in restriction_check['violations']:
                    conflicts.append({
                        'type': 'equipment',
                        'severity': violation.get('severity', 'error'),
                        'description': violation.get('reason'),
                        'job_ids': [job.id],
                        'user_ids': violation.get('user_id', [])
                    })

        # Dependency check
        for dependency in job.dependencies:
            depends_on = dependency.depends_on_job
            if depends_on.day.date > day.date:
                conflicts.append({
                    'type': 'dependency',
                    'severity': 'error',
                    'description': f'Job {job.id} scheduled before its dependency (job {depends_on.id})',
                    'job_ids': [job.id, depends_on.id],
                    'user_ids': []
                })
            elif depends_on.day.date == day.date:
                # Same day - check time order if time slots exist
                if job.start_time and depends_on.end_time:
                    if job.start_time < depends_on.end_time:
                        conflicts.append({
                            'type': 'dependency',
                            'severity': 'warning',
                            'description': f'Job {job.id} may start before dependency (job {depends_on.id}) completes',
                            'job_ids': [job.id, depends_on.id],
                            'user_ids': []
                        })

        return conflicts

    @staticmethod
    def detect_job_conflicts(job_id: int) -> list:
        """Detect conflicts for a specific job.

        Args:
            job_id: ID of the job

        Returns:
            List of detected conflicts
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        conflicts = WorkPlanService._detect_job_conflicts_internal(job, job.day, {}, {})
        return conflicts

    @staticmethod
    def resolve_conflict(conflict_id: int, resolution: str,
                         resolved_by_id: int) -> SchedulingConflict:
        """Mark a conflict as resolved.

        Args:
            conflict_id: ID of the conflict
            resolution: Resolution description
            resolved_by_id: ID of user resolving the conflict

        Returns:
            Updated SchedulingConflict instance
        """
        conflict = db.session.get(SchedulingConflict, conflict_id)
        if not conflict:
            raise NotFoundError(f"Conflict with ID {conflict_id} not found")

        conflict.resolution = resolution
        conflict.resolved_at = datetime.utcnow()
        conflict.resolved_by_id = resolved_by_id

        db.session.commit()

        logger.info(f"Resolved conflict {conflict_id}")
        return conflict

    @staticmethod
    def ignore_conflict(conflict_id: int, ignored_by_id: int) -> SchedulingConflict:
        """Mark a conflict as ignored.

        Args:
            conflict_id: ID of the conflict
            ignored_by_id: ID of user ignoring the conflict

        Returns:
            Updated SchedulingConflict instance
        """
        conflict = db.session.get(SchedulingConflict, conflict_id)
        if not conflict:
            raise NotFoundError(f"Conflict with ID {conflict_id} not found")

        conflict.is_ignored = True
        conflict.resolved_by_id = ignored_by_id
        conflict.resolved_at = datetime.utcnow()
        conflict.resolution = 'Acknowledged and ignored'

        db.session.commit()

        logger.info(f"Ignored conflict {conflict_id}")
        return conflict

    # =========================================================================
    # VALIDATION
    # =========================================================================

    @staticmethod
    def validate_plan(plan_id: int) -> dict:
        """Full plan validation.

        Args:
            plan_id: ID of the work plan

        Returns:
            Dictionary with valid status, errors, and warnings
        """
        plan = db.session.get(WorkPlan, plan_id)
        if not plan:
            raise NotFoundError(f"Work plan with ID {plan_id} not found")

        errors = []
        warnings = []

        # Check for empty days
        for day in plan.days:
            if len(day.jobs) == 0:
                warnings.append({
                    'type': 'empty_day',
                    'message': f'No jobs scheduled for {day.date}',
                    'day_id': day.id
                })

        # Check for unassigned jobs
        for day in plan.days:
            for job in day.jobs:
                if len(job.assignments) == 0:
                    errors.append({
                        'type': 'unassigned_job',
                        'message': f'Job {job.id} has no assigned workers',
                        'job_id': job.id
                    })

        # Detect scheduling conflicts
        conflicts = WorkPlanService.detect_conflicts(plan_id)
        for conflict in conflicts:
            if conflict['severity'] == 'error':
                errors.append({
                    'type': conflict['type'],
                    'message': conflict['description'],
                    'job_ids': conflict.get('job_ids')
                })
            else:
                warnings.append({
                    'type': conflict['type'],
                    'message': conflict['description'],
                    'job_ids': conflict.get('job_ids')
                })

        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        }

    @staticmethod
    def validate_job_assignment(job_id: int, user_id: int) -> dict:
        """Validate if user can be assigned to job.

        Checks: capacity, skills, equipment restrictions, leave status

        Args:
            job_id: ID of the job
            user_id: ID of the user

        Returns:
            Dictionary with can_assign status and issues
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            raise NotFoundError(f"Job with ID {job_id} not found")

        user = db.session.get(User, user_id)
        if not user:
            raise NotFoundError(f"User with ID {user_id} not found")

        issues = []

        # Check if user is active
        if not user.is_active:
            issues.append({
                'type': 'inactive_user',
                'message': 'User is not active',
                'severity': 'error'
            })

        # Check leave status
        if user.is_on_leave:
            issues.append({
                'type': 'on_leave',
                'message': 'User is currently on leave',
                'severity': 'error'
            })

        # Check for approved leave on job date
        leave = Leave.query.filter(
            Leave.user_id == user_id,
            Leave.status == 'approved',
            Leave.date_from <= job.day.date,
            Leave.date_to >= job.day.date
        ).first()
        if leave:
            issues.append({
                'type': 'leave_scheduled',
                'message': f'User has approved leave on {job.day.date}',
                'severity': 'error'
            })

        # Check capacity
        capacity = WorkPlanService.check_capacity_violation(
            user_id, job.day.id, job.estimated_hours
        )
        if capacity['violated']:
            issues.append({
                'type': 'capacity_violation',
                'message': f'Assignment would exceed capacity ({capacity["total_hours"]} hours)',
                'severity': 'warning'
            })

        # Check skills
        skill_check = WorkPlanService.check_skill_requirements(job_id, user_id)
        if not skill_check['meets_requirements']:
            issues.append({
                'type': 'missing_skills',
                'message': f'Missing required skills: {", ".join(skill_check["missing_skills"])}',
                'severity': 'error'
            })

        # Check equipment restrictions
        if job.equipment_id:
            restriction_check = WorkPlanService.check_equipment_restrictions(
                job.equipment_id, job.day.date, [user_id]
            )
            for violation in restriction_check['violations']:
                issues.append({
                    'type': 'equipment_restriction',
                    'message': violation.get('reason'),
                    'severity': violation.get('severity', 'warning')
                })

        can_assign = all(
            issue['severity'] != 'error'
            for issue in issues
        )

        return {
            'can_assign': can_assign,
            'issues': issues
        }

    # =========================================================================
    # NOTIFICATIONS
    # =========================================================================

    @staticmethod
    def notify_job_changes(job_id: int, change_type: str, changed_by_id: int):
        """Send notifications for job changes.

        Args:
            job_id: ID of the job
            change_type: Type of change (created, updated, deleted, moved)
            changed_by_id: ID of user who made the change
        """
        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            return

        # Get all assigned users except the one who made the change
        assigned_user_ids = [
            a.user_id for a in job.assignments
            if a.user_id != changed_by_id
        ]

        if not assigned_user_ids:
            return

        title = f"Job {change_type.capitalize()}"
        message = f"A job assigned to you has been {change_type}"

        if job.equipment:
            message += f" ({job.equipment.name})"

        NotificationService.create_bulk_notification(
            user_ids=assigned_user_ids,
            type='job_updated',
            title=title,
            message=message,
            related_type='work_plan_job',
            related_id=job_id,
            priority='info'
        )

    @staticmethod
    def notify_assignment_changes(job_id: int, user_id: int, action: str,
                                  notified_by_id: int):
        """Send notification when user is assigned/unassigned.

        Args:
            job_id: ID of the job
            user_id: ID of the user affected
            action: Action (assigned, unassigned)
            notified_by_id: ID of user who made the change
        """
        if user_id == notified_by_id:
            return

        job = db.session.get(WorkPlanJob, job_id)
        if not job:
            return

        if action == 'assigned':
            title = "New Job Assignment"
            message = f"You have been assigned to a {job.job_type} job"
            priority = 'info'
        else:
            title = "Job Unassignment"
            message = f"You have been removed from a {job.job_type} job"
            priority = 'warning'

        if job.equipment:
            message += f" for {job.equipment.name}"

        message += f" on {job.day.date.strftime('%B %d, %Y')}"

        NotificationService.create_notification(
            user_id=user_id,
            type='job_assigned' if action == 'assigned' else 'job_unassigned',
            title=title,
            message=message,
            related_type='work_plan_job',
            related_id=job_id,
            priority=priority
        )

    # =========================================================================
    # REPORTS
    # =========================================================================

    @staticmethod
    def get_performance_report(user_id: int = None, period: str = 'weekly') -> dict:
        """Get performance metrics.

        Args:
            user_id: Specific user ID (optional, None for all)
            period: Report period (daily, weekly, monthly)

        Returns:
            Dictionary with performance metrics
        """
        from app.models import WorkPlanJobTracking

        # Determine date range
        today = datetime.utcnow().date()
        if period == 'daily':
            start_date = today
        elif period == 'weekly':
            start_date = today - timedelta(days=7)
        else:  # monthly
            start_date = today - timedelta(days=30)

        # Build query
        query = db.session.query(
            WorkPlanJob.id,
            WorkPlanJob.job_type,
            WorkPlanJob.estimated_hours,
            WorkPlanJobTracking.actual_duration_minutes
        ).outerjoin(
            WorkPlanJobTracking,
            WorkPlanJobTracking.work_plan_job_id == WorkPlanJob.id
        ).join(
            WorkPlanDay,
            WorkPlanDay.id == WorkPlanJob.work_plan_day_id
        ).filter(
            WorkPlanDay.date >= start_date
        )

        if user_id:
            query = query.join(
                WorkPlanAssignment,
                WorkPlanAssignment.work_plan_job_id == WorkPlanJob.id
            ).filter(
                WorkPlanAssignment.user_id == user_id
            )

        results = query.all()

        # Calculate metrics
        total_jobs = len(results)
        completed_jobs = sum(1 for r in results if r.actual_duration_minutes is not None)
        total_estimated = sum(r.estimated_hours or 0 for r in results)
        total_actual = sum((r.actual_duration_minutes or 0) / 60 for r in results)

        by_type = {}
        for r in results:
            if r.job_type not in by_type:
                by_type[r.job_type] = {'count': 0, 'estimated': 0, 'actual': 0}
            by_type[r.job_type]['count'] += 1
            by_type[r.job_type]['estimated'] += r.estimated_hours or 0
            by_type[r.job_type]['actual'] += (r.actual_duration_minutes or 0) / 60

        accuracy = (total_estimated / total_actual * 100) if total_actual > 0 else 0

        return {
            'period': period,
            'start_date': start_date.isoformat(),
            'end_date': today.isoformat(),
            'total_jobs': total_jobs,
            'completed_jobs': completed_jobs,
            'completion_rate': (completed_jobs / total_jobs * 100) if total_jobs > 0 else 0,
            'total_estimated_hours': round(total_estimated, 2),
            'total_actual_hours': round(total_actual, 2),
            'estimation_accuracy': round(accuracy, 1),
            'by_job_type': by_type
        }

    @staticmethod
    def get_completion_report(date_from: date, date_to: date) -> dict:
        """Get completion statistics.

        Args:
            date_from: Start date
            date_to: End date

        Returns:
            Dictionary with completion statistics
        """
        from app.models import WorkPlanJobTracking

        # Get jobs in date range
        jobs = WorkPlanJob.query.join(WorkPlanDay).filter(
            WorkPlanDay.date >= date_from,
            WorkPlanDay.date <= date_to
        ).all()

        total = len(jobs)
        completed = 0
        on_time = 0
        delayed = 0

        by_day = {}

        for job in jobs:
            day_str = job.day.date.isoformat()
            if day_str not in by_day:
                by_day[day_str] = {'total': 0, 'completed': 0}
            by_day[day_str]['total'] += 1

            tracking = WorkPlanJobTracking.query.filter_by(
                work_plan_job_id=job.id
            ).first()

            if tracking and tracking.status == 'completed':
                completed += 1
                by_day[day_str]['completed'] += 1

                # Check if on time
                if tracking.actual_duration_minutes and job.estimated_hours:
                    estimated_minutes = job.estimated_hours * 60
                    if tracking.actual_duration_minutes <= estimated_minutes * 1.1:  # 10% buffer
                        on_time += 1
                    else:
                        delayed += 1

        return {
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'total_jobs': total,
            'completed_jobs': completed,
            'completion_rate': round((completed / total * 100) if total > 0 else 0, 1),
            'on_time': on_time,
            'delayed': delayed,
            'by_day': by_day
        }

    @staticmethod
    def get_time_accuracy_report() -> list:
        """Get time estimation accuracy by job type/equipment.

        Returns:
            List of accuracy metrics by job type and equipment
        """
        from app.models import WorkPlanJobTracking

        # Query jobs with tracking data
        results = db.session.query(
            WorkPlanJob.job_type,
            WorkPlanJob.equipment_id,
            Equipment.name.label('equipment_name'),
            func.count(WorkPlanJob.id).label('job_count'),
            func.sum(WorkPlanJob.estimated_hours).label('total_estimated'),
            func.sum(WorkPlanJobTracking.actual_duration_minutes).label('total_actual_minutes')
        ).outerjoin(
            WorkPlanJobTracking,
            WorkPlanJobTracking.work_plan_job_id == WorkPlanJob.id
        ).outerjoin(
            Equipment,
            Equipment.id == WorkPlanJob.equipment_id
        ).filter(
            WorkPlanJobTracking.actual_duration_minutes.isnot(None)
        ).group_by(
            WorkPlanJob.job_type,
            WorkPlanJob.equipment_id,
            Equipment.name
        ).all()

        report = []
        for r in results:
            actual_hours = (r.total_actual_minutes or 0) / 60
            estimated_hours = r.total_estimated or 0

            if actual_hours > 0:
                accuracy = (estimated_hours / actual_hours) * 100
            else:
                accuracy = 0

            report.append({
                'job_type': r.job_type,
                'equipment_id': r.equipment_id,
                'equipment_name': r.equipment_name,
                'job_count': r.job_count,
                'total_estimated_hours': round(estimated_hours, 2),
                'total_actual_hours': round(actual_hours, 2),
                'accuracy_percentage': round(accuracy, 1),
                'variance_hours': round(estimated_hours - actual_hours, 2)
            })

        return sorted(report, key=lambda x: abs(x['variance_hours']), reverse=True)

    @staticmethod
    def export_plan_to_excel(plan_id: int) -> bytes:
        """Export plan to Excel format.

        Args:
            plan_id: ID of the work plan

        Returns:
            Excel file as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise BusinessError("openpyxl is required for Excel export")

        plan = db.session.get(WorkPlan, plan_id)
        if not plan:
            raise NotFoundError(f"Work plan with ID {plan_id} not found")

        wb = Workbook()
        ws = wb.active
        ws.title = "Work Plan"

        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = f"Work Plan: {plan.week_start} to {plan.week_end}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')

        # Headers
        headers = ['Date', 'Day', 'Job Type', 'Equipment', 'SAP Order',
                   'Est. Hours', 'Priority', 'Assigned Workers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Data
        row = 4
        for day in plan.days:
            for job in day.jobs:
                ws.cell(row=row, column=1, value=day.date.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=2, value=day.date.strftime('%A')).border = border
                ws.cell(row=row, column=3, value=job.job_type.upper()).border = border
                ws.cell(row=row, column=4, value=job.equipment.name if job.equipment else '-').border = border
                ws.cell(row=row, column=5, value=job.sap_order_number or '-').border = border
                ws.cell(row=row, column=6, value=job.estimated_hours).border = border
                ws.cell(row=row, column=7, value=job.priority.capitalize()).border = border

                workers = ', '.join(a.user.full_name for a in job.assignments if a.user)
                ws.cell(row=row, column=8, value=workers or '-').border = border

                row += 1

        # Adjust column widths
        column_widths = [12, 10, 12, 25, 15, 10, 10, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
