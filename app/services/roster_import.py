"""Read the team roster out of the engineering workbook the courier delivers.

The file is `Engineering 2026_v1.xlsm` and the sheet is `All Employees Off Days`.
It arrived every day from 2026-08-28 onward, was stored, and was read by nothing
— the app's existing roster upload could not have read it even by hand, for two
reasons: it rejects `.xlsm` on the filename, and it parses a completely
different layout.

WHAT THE SHEET ACTUALLY LOOKS LIKE (measured, not assumed):

    row 1   Company | Section | Name | SAP ID | Section | Sub Section | Post |  | January
    row 2                                                                      | 2026-01-01 | ...
    row 4+  BGT     | HSE     | ...  | 500000 | ...                            | X | D | ...

  * SAP ID is column D (index 3), not column A.
  * The dates are on ROW 2, from column I (index 8) onward, as real datetimes.
  * 381 columns, of which the last few are totals — their row-2 cell is NOT a
    date, and that is exactly how they are told apart. Reading them as shifts is
    where the stray 14 / 76 / 77 values came from.
  * 103 employees.

ALI'S RULES, 2026-09-04:

  * "what i change in the app should be kept as what my change is" — an import
    never overwrites a day a person set. See RosterEntry.source.
  * "applied as it is the first apply for it" — the first import lands whole.
    That needs no flag: rows from the old upload path carry source NULL, which
    is replaceable, and nothing can be marked manual until someone marks it.
  * "many employees are not in the app ... just drop them" — an unmatched SAP ID
    is skipped, not an error. It is COUNTED, because silently dropping a third
    of the sheet is how "just drop them" turns into "where did my people go".
"""

import io
import logging
from datetime import date, datetime

logger = logging.getLogger(__name__)

SHEET_NAME = 'All Employees Off Days'

# Row and column geography, 0-based, measured from the real file.
HEADER_ROW = 0        # Company | Section | Name | SAP ID | ...
DATE_ROW = 1          # the datetimes live here
FIRST_DATA_ROW = 3
SAP_ID_COL = 3
FIRST_DATE_COL = 8

# The codes Ali confirmed. COM is a holiday.
#
# Everything that is not D or N means the man cannot be given work that day, so
# an unknown code is skipped rather than guessed: mis-reading a code as 'day'
# would put a man on shift who is at home, while skipping only loses capacity.
# Skipped codes are counted, so a new code appearing in the sheet is visible
# rather than silent.
#
# The workbook's own `Instruction` sheet is the authority if one of these ever
# looks wrong.
SHIFT_CODES = {
    'D': 'day',
    'N': 'night',
    'X': 'off',
    'PL': 'leave',    # planned / paid leave
    'PSL': 'leave',   # paid sick leave
    'COM': 'leave',   # holiday (Ali, 2026-09-04)
    'EX': 'leave',    # leave earned by working extra (Ali, 2026-09-04).
                      # 'Extra days' reads like extra work, and mapping it to a
                      # shift would have put a man on the board on a day he had
                      # earned off. It is leave: he is not at work.
}

# An import may replace a day it wrote itself, or one from before this existed.
# It must never replace 'manual' or 'swap'.
REPLACEABLE_SOURCES = (None, 'import')


def _sap_id(value):
    """The sheet's SAP ID as the app stores it.

    openpyxl hands a numeric cell back as a float, so 500000 arrives as
    500000.0 and would match nobody — which under Ali's "just drop them" rule
    would silently discard the entire sheet.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text:
        return None
    if text.endswith('.0'):
        head = text[:-2]
        if head.isdigit():
            return head
    return text


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse(workbook_bytes):
    """(rows, report) — one row per person, without touching the database.

    A row is {'sap_id': str, 'name': str, 'days': {date: shift}}.
    """
    import openpyxl

    report = {'sheet': SHEET_NAME, 'people_in_sheet': 0, 'date_columns': 0,
              'unknown_codes': {}, 'blank_cells_skipped': 0,
              'duplicate_sap_ids': {}}

    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes),
                                      data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        # Never fall back to wb.active. That is what "works today and breaks the
        # day somebody saves the file on a different tab" looks like.
        raise ValueError(
            f'The workbook has no sheet named {SHEET_NAME!r}. '
            f'It has: {workbook.sheetnames}')

    grid = list(workbook[SHEET_NAME].iter_rows(values_only=True))
    if len(grid) <= FIRST_DATA_ROW:
        return [], report

    # A column is a DATE column only if row 2 holds a real date. The trailing
    # totals columns fail that test, which is how they stay out of the roster.
    date_row = grid[DATE_ROW]
    date_columns = []
    for index in range(FIRST_DATE_COL, len(date_row)):
        day = _as_date(date_row[index])
        if day:
            date_columns.append((index, day))
    report['date_columns'] = len(date_columns)

    unknown = {}
    duplicates = {}
    seen_ids = {}
    rows = []
    for line in grid[FIRST_DATA_ROW:]:
        if len(line) <= SAP_ID_COL:
            continue
        sap_id = _sap_id(line[SAP_ID_COL])
        if not sap_id:
            continue
        report['people_in_sheet'] += 1

        days = {}
        for index, day in date_columns:
            if index >= len(line):
                continue
            raw = line[index]
            if raw is None or str(raw).strip() == '':
                report['blank_cells_skipped'] += 1
                continue
            code = str(raw).strip().upper()
            shift = SHIFT_CODES.get(code)
            if shift is None:
                unknown[code] = unknown.get(code, 0) + 1
                continue
            days[day] = shift

        # MERGED, not appended. Two sheet lines with the same SAP ID both
        # resolve to one user, and roster_entries has
        # UniqueConstraint('user_id','date') — so appending them added two rows
        # for one day and the COMMIT raised IntegrityError, losing the entire
        # import. Not once: the duplicate survives every daily save of the
        # workbook, and the scheduler records its claim BEFORE the work, so the
        # file would never be retried. 106 rows carry 103 SAP ids.
        #
        # The later line wins, matching how a person reading down the sheet
        # would take the last thing written about someone.
        if sap_id in seen_ids:
            merged = rows[seen_ids[sap_id]]
            merged['days'].update(days)
            duplicates[sap_id] = duplicates.get(sap_id, 1) + 1
            continue
        seen_ids[sap_id] = len(rows)
        rows.append({'sap_id': sap_id,
                     'name': (line[2] or '') if len(line) > 2 else '',
                     'days': days})

    report['unknown_codes'] = dict(sorted(unknown.items(),
                                          key=lambda kv: -kv[1])[:20])
    report['duplicate_sap_ids'] = duplicates
    return rows, report


def apply_roster(workbook_bytes, dry_run=False, commit=True):
    """Write the sheet into roster_entries. Returns a report.

    Writes nothing when `dry_run` is set, which is how a change is inspected
    before it touches a live roster.
    """
    from app.extensions import db
    from app.models import User
    from app.models.roster import RosterEntry

    rows, report = parse(workbook_bytes)
    report.update({'dry_run': dry_run, 'matched_people': 0, 'dropped_people': 0,
                   'dropped_sap_ids': [], 'created': 0, 'updated': 0,
                   'left_alone_manual': 0})

    if not rows:
        return report

    # One query, not one per person: 103 lookups against a free-plan database is
    # a minute of round trips for a screen's worth of data.
    wanted = {row['sap_id'] for row in rows}
    users = {u.sap_id: u.id for u in
             User.query.filter(User.sap_id.in_(wanted)).all()}

    matched = [row for row in rows if row['sap_id'] in users]
    for row in rows:
        if row['sap_id'] not in users:
            report['dropped_people'] += 1
            # Named, not just counted. "Just drop them" is right, but a silent
            # drop of a third of the sheet is how people go missing unnoticed.
            if len(report['dropped_sap_ids']) < 50:
                report['dropped_sap_ids'].append(row['sap_id'])
    report['matched_people'] = len(matched)

    if not matched:
        return report

    user_ids = [users[row['sap_id']] for row in matched]
    existing = {}
    for entry in RosterEntry.query.filter(RosterEntry.user_id.in_(user_ids)).all():
        existing[(entry.user_id, entry.date)] = entry

    for row in matched:
        user_id = users[row['sap_id']]
        for day, shift in row['days'].items():
            entry = existing.get((user_id, day))
            if entry is None:
                if not dry_run:
                    db.session.add(RosterEntry(user_id=user_id, date=day,
                                               shift=shift, source='import'))
                report['created'] += 1
                continue
            if entry.source not in REPLACEABLE_SOURCES:
                # A person set this day. Ali: keep it as his change.
                report['left_alone_manual'] += 1
                continue
            if entry.shift != shift or entry.source != 'import':
                if not dry_run:
                    entry.shift = shift
                    entry.source = 'import'
                report['updated'] += 1

    # `commit=False` lets the caller put the roster rows and the file's
    # parsed_at stamp in ONE transaction, so a crash rolls back both and the
    # file is retried rather than marked done and lost.
    if not dry_run and commit:
        db.session.commit()

    logger.info('Roster import: %s', {k: v for k, v in report.items()
                                      if k != 'dropped_sap_ids'})
    return report
