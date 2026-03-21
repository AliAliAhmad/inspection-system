"""
Materials management endpoints.
Handles stock materials, material kits, and consumption tracking.
"""

from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.extensions import db
from app.models import Material, MaterialKit, MaterialKitItem, User
from app.exceptions.api_exceptions import ValidationError, NotFoundError, ForbiddenError
from app.utils.decorators import get_current_user
from datetime import datetime, date
from app.services.material_ai_service import MaterialAIService
from app.services.stock_alert_service import StockAlertService
from app.services.material_service import MaterialService

bp = Blueprint('materials', __name__)

material_ai = MaterialAIService()
stock_alerts = StockAlertService()


def engineer_or_admin_required():
    """Check if user is engineer or admin."""
    user = get_current_user()
    if user.role not in ['admin', 'engineer', 'quality_engineer']:
        raise ForbiddenError("Only engineers and admins can access this resource")
    return user


# ==================== MATERIALS ====================

@bp.route('', methods=['GET'])
@jwt_required()
def list_materials():
    """
    List all materials with optional filtering.

    Query params:
        - category: Filter by category
        - low_stock: If true, only show low stock items
        - search: Search by code or name
        - active_only: If true, only show active items (default true)
    """
    category = request.args.get('category')
    low_stock = request.args.get('low_stock', 'false').lower() == 'true'
    search = request.args.get('search', '').strip()
    active_only = request.args.get('active_only', 'true').lower() == 'true'

    user = get_current_user()
    language = user.language or 'en'

    query = Material.query

    if active_only:
        query = query.filter(Material.is_active == True)

    if category:
        query = query.filter(Material.category == category)

    if search:
        search_term = f'%{search}%'
        query = query.filter(
            db.or_(
                Material.code.ilike(search_term),
                Material.name.ilike(search_term),
                Material.name_ar.ilike(search_term)
            )
        )

    materials = query.order_by(Material.name).all()

    # Filter low stock if requested
    if low_stock:
        materials = [m for m in materials if m.is_low_stock()]

    return jsonify({
        'status': 'success',
        'materials': [m.to_dict(language) for m in materials],
        'count': len(materials)
    }), 200


@bp.route('/<int:material_id>', methods=['GET'])
@jwt_required()
def get_material(material_id):
    """Get a single material by ID."""
    material = db.session.get(Material, material_id)
    if not material:
        raise NotFoundError("Material not found")

    user = get_current_user()
    language = user.language or 'en'

    return jsonify({
        'status': 'success',
        'material': material.to_dict(language)
    }), 200


@bp.route('', methods=['POST'])
@jwt_required()
def create_material():
    """
    Create a new material. Engineers and admins only.

    Request body:
        {
            "code": "MAT-001",
            "name": "Hydraulic Oil",
            "name_ar": "زيت هيدروليكي",
            "category": "lubricant",
            "unit": "liters",
            "current_stock": 100,
            "min_stock": 20
        }
    """
    user = engineer_or_admin_required()
    data = request.get_json()

    if not data:
        raise ValidationError("Request body is required")

    required = ['code', 'name', 'category', 'unit']
    missing = [f for f in required if not data.get(f)]
    if missing:
        raise ValidationError(f"Missing required fields: {', '.join(missing)}")

    # Check for duplicate code
    existing = Material.query.filter_by(code=data['code']).first()
    if existing:
        raise ValidationError(f"Material with code {data['code']} already exists")

    material = Material(
        code=data['code'],
        name=data['name'],
        name_ar=data.get('name_ar'),
        category=data['category'],
        unit=data['unit'],
        current_stock=data.get('current_stock', 0),
        min_stock=data.get('min_stock', 0),
        consumption_start_date=datetime.utcnow().date()
    )

    db.session.add(material)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Material created',
        'material': material.to_dict(user.language or 'en')
    }), 201


@bp.route('/<int:material_id>', methods=['PUT'])
@jwt_required()
def update_material(material_id):
    """Update a material. Engineers and admins only."""
    user = engineer_or_admin_required()

    material = db.session.get(Material, material_id)
    if not material:
        raise NotFoundError("Material not found")

    data = request.get_json()
    if not data:
        raise ValidationError("Request body is required")

    # Update fields
    if 'name' in data:
        material.name = data['name']
    if 'name_ar' in data:
        material.name_ar = data['name_ar']
    if 'category' in data:
        material.category = data['category']
    if 'unit' in data:
        material.unit = data['unit']
    if 'current_stock' in data:
        material.current_stock = data['current_stock']
    if 'min_stock' in data:
        material.min_stock = data['min_stock']
    if 'is_active' in data:
        material.is_active = data['is_active']

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Material updated',
        'material': material.to_dict(user.language or 'en')
    }), 200


# ==================== STOCK OPERATIONS ====================

@bp.route('/<int:material_id>/consume', methods=['POST'])
@jwt_required()
def consume_material(material_id):
    """
    Consume material with full tracking.
    Body: { quantity, reason, job_id?, batch_id? }
    """
    user = get_current_user()
    data = request.get_json()
    result = MaterialService.consume(
        material_id=material_id,
        quantity=data['quantity'],
        user_id=user.id,
        reason=data.get('reason'),
        job_id=data.get('job_id'),
        batch_id=data.get('batch_id')
    )
    return jsonify({'status': 'success', 'data': result})


@bp.route('/<int:material_id>/restock', methods=['POST'])
@jwt_required()
def restock_material(material_id):
    """
    Restock material with batch creation.
    Body: { quantity, batch_info?, vendor_id?, unit_price? }
    """
    user = engineer_or_admin_required()
    data = request.get_json()
    result = MaterialService.restock(
        material_id=material_id,
        quantity=data['quantity'],
        user_id=user.id,
        batch_info=data.get('batch_info'),
        vendor_id=data.get('vendor_id')
    )
    return jsonify({'status': 'success', 'data': result})


@bp.route('/<int:material_id>/adjust', methods=['POST'])
@jwt_required()
def adjust_stock(material_id):
    """
    Manual stock adjustment with audit trail.
    Body: { new_quantity, reason }
    """
    user = engineer_or_admin_required()
    data = request.get_json()
    result = MaterialService.adjust(
        material_id=material_id,
        new_quantity=data['new_quantity'],
        user_id=user.id,
        reason=data['reason']
    )
    return jsonify({'status': 'success', 'data': result})


@bp.route('/<int:material_id>/transfer', methods=['POST'])
@jwt_required()
def transfer_stock(material_id):
    """
    Transfer between locations.
    Body: { from_location_id, to_location_id, quantity }
    """
    user = engineer_or_admin_required()
    data = request.get_json()
    result = MaterialService.transfer(
        material_id=material_id,
        from_location_id=data['from_location_id'],
        to_location_id=data['to_location_id'],
        quantity=data['quantity'],
        user_id=user.id
    )
    return jsonify({'status': 'success', 'data': result})


@bp.route('/<int:material_id>/reserve', methods=['POST'])
@jwt_required()
def reserve_stock(material_id):
    """
    Reserve stock for upcoming job.
    Body: { quantity, job_id?, work_plan_id?, needed_by_date? }
    """
    user = get_current_user()
    data = request.get_json()
    result = MaterialService.reserve(
        material_id=material_id,
        quantity=data['quantity'],
        user_id=user.id,
        job_id=data.get('job_id'),
        needed_by=data.get('needed_by_date')
    )
    return jsonify({'status': 'success', 'data': result})


@bp.route('/reservations/<int:reservation_id>/fulfill', methods=['POST'])
@jwt_required()
def fulfill_reservation(reservation_id):
    """Fulfill a reservation."""
    user = get_current_user()
    result = MaterialService.fulfill_reservation(reservation_id, user.id)
    return jsonify({'status': 'success', 'data': result})


@bp.route('/reservations/<int:reservation_id>/cancel', methods=['POST'])
@jwt_required()
def cancel_reservation(reservation_id):
    """Cancel a reservation."""
    user = get_current_user()
    data = request.get_json() or {}
    result = MaterialService.cancel_reservation(reservation_id, user.id, data.get('reason'))
    return jsonify({'status': 'success', 'data': result})


@bp.route('/<int:material_id>/summary', methods=['GET'])
@jwt_required()
def get_stock_summary(material_id):
    """Get comprehensive stock summary."""
    summary = MaterialService.get_stock_summary(material_id)
    return jsonify({'status': 'success', 'data': summary})


@bp.route('/stock-check', methods=['POST'])
@jwt_required()
def check_low_stock():
    """
    Check for low stock materials and return warnings.
    Engineers and admins only.
    """
    user = engineer_or_admin_required()
    language = user.language or 'en'

    # Get all active materials with low stock
    materials = Material.query.filter(Material.is_active == True).all()
    low_stock_items = [m for m in materials if m.is_low_stock()]

    # TODO: Send notifications to store/engineers

    return jsonify({
        'status': 'success',
        'low_stock_count': len(low_stock_items),
        'low_stock_materials': [m.to_dict(language) for m in low_stock_items]
    }), 200


# ==================== MATERIAL KITS ====================

@bp.route('/kits', methods=['GET'])
@jwt_required()
def list_kits():
    """List all material kits."""
    user = get_current_user()
    language = user.language or 'en'

    equipment_type = request.args.get('equipment_type')
    active_only = request.args.get('active_only', 'true').lower() == 'true'

    query = MaterialKit.query

    if active_only:
        query = query.filter(MaterialKit.is_active == True)

    if equipment_type:
        query = query.filter(
            db.or_(
                MaterialKit.equipment_type == equipment_type,
                MaterialKit.equipment_type.is_(None)
            )
        )

    kits = query.order_by(MaterialKit.name).all()

    return jsonify({
        'status': 'success',
        'kits': [k.to_dict(language) for k in kits],
        'count': len(kits)
    }), 200


@bp.route('/kits/<int:kit_id>', methods=['GET'])
@jwt_required()
def get_kit(kit_id):
    """Get a single kit by ID."""
    kit = db.session.get(MaterialKit, kit_id)
    if not kit:
        raise NotFoundError("Material kit not found")

    user = get_current_user()
    language = user.language or 'en'

    return jsonify({
        'status': 'success',
        'kit': kit.to_dict(language)
    }), 200


@bp.route('/kits', methods=['POST'])
@jwt_required()
def create_kit():
    """
    Create a new material kit. Engineers and admins only.

    Request body:
        {
            "name": "Crane 2000 PM Kit",
            "name_ar": "طقم صيانة رافعة 2000",
            "description": "Standard PM kit for Crane 2000",
            "equipment_type": "STS Crane",
            "items": [
                {"material_id": 1, "quantity": 2},
                {"material_id": 3, "quantity": 1}
            ]
        }
    """
    user = engineer_or_admin_required()
    data = request.get_json()

    if not data or not data.get('name'):
        raise ValidationError("Kit name is required")

    kit = MaterialKit(
        name=data['name'],
        name_ar=data.get('name_ar'),
        description=data.get('description'),
        equipment_type=data.get('equipment_type')
    )

    db.session.add(kit)
    db.session.flush()  # Get kit ID

    # Add items
    for item_data in data.get('items', []):
        material_id = item_data.get('material_id')
        quantity = item_data.get('quantity', 1)

        if not material_id:
            continue

        material = db.session.get(Material, material_id)
        if not material:
            raise ValidationError(f"Material {material_id} not found")

        item = MaterialKitItem(
            kit_id=kit.id,
            material_id=material_id,
            quantity=quantity
        )
        db.session.add(item)

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Material kit created',
        'kit': kit.to_dict(user.language or 'en')
    }), 201


@bp.route('/kits/<int:kit_id>', methods=['PUT'])
@jwt_required()
def update_kit(kit_id):
    """Update a material kit. Engineers and admins only."""
    user = engineer_or_admin_required()

    kit = db.session.get(MaterialKit, kit_id)
    if not kit:
        raise NotFoundError("Material kit not found")

    data = request.get_json()
    if not data:
        raise ValidationError("Request body is required")

    # Update fields
    if 'name' in data:
        kit.name = data['name']
    if 'name_ar' in data:
        kit.name_ar = data['name_ar']
    if 'description' in data:
        kit.description = data['description']
    if 'equipment_type' in data:
        kit.equipment_type = data['equipment_type']
    if 'is_active' in data:
        kit.is_active = data['is_active']

    # Update items if provided
    if 'items' in data:
        # Remove existing items
        MaterialKitItem.query.filter_by(kit_id=kit.id).delete()

        # Add new items
        for item_data in data['items']:
            material_id = item_data.get('material_id')
            quantity = item_data.get('quantity', 1)

            if not material_id:
                continue

            material = db.session.get(Material, material_id)
            if not material:
                raise ValidationError(f"Material {material_id} not found")

            item = MaterialKitItem(
                kit_id=kit.id,
                material_id=material_id,
                quantity=quantity
            )
            db.session.add(item)

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Material kit updated',
        'kit': kit.to_dict(user.language or 'en')
    }), 200


@bp.route('/kits/<int:kit_id>', methods=['DELETE'])
@jwt_required()
def delete_kit(kit_id):
    """Delete a material kit. Engineers and admins only."""
    user = engineer_or_admin_required()

    kit = db.session.get(MaterialKit, kit_id)
    if not kit:
        raise NotFoundError("Material kit not found")

    db.session.delete(kit)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Material kit deleted'
    }), 200


@bp.route('/import', methods=['POST'])
@jwt_required()
def import_materials():
    """
    Import materials from Excel file. Engineers and admins only.

    Expected columns: code, name, name_ar, category, unit, current_stock, min_stock
    """
    import pandas as pd
    from io import BytesIO

    user = engineer_or_admin_required()

    if 'file' not in request.files:
        raise ValidationError("No file uploaded")

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise ValidationError("File must be Excel format (.xlsx or .xls)")

    try:
        df = pd.read_excel(BytesIO(file.read()))
    except Exception as e:
        raise ValidationError(f"Failed to read Excel file: {str(e)}")

    required_columns = ['code', 'name', 'category', 'unit']
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        raise ValidationError(f"Missing required columns: {', '.join(missing)}")

    # Valid categories per DB constraint
    VALID_CATEGORIES = {'lubricant', 'filter', 'spare_part', 'consumable', 'electrical', 'mechanical', 'hvac', 'other'}
    # Map common Excel variations to valid DB values
    CATEGORY_MAP = {
        'spare parts': 'spare_part', 'spare part': 'spare_part', 'spares': 'spare_part',
        'filters': 'filter', 'oil filter': 'filter', 'air filter': 'filter',
        'lubricants': 'lubricant', 'oil': 'lubricant', 'grease': 'lubricant',
        'consumables': 'consumable',
        'electric': 'electrical', 'electrics': 'electrical',
        'mech': 'mechanical',
    }

    def safe_str(val, default=None):
        """Convert pandas value to string, treating NaN/nan as None."""
        if pd.isna(val):
            return default
        s = str(val).strip()
        if s.lower() in ('nan', 'none', ''):
            return default
        return s

    def normalize_category(raw):
        """Map Excel category to valid DB category."""
        if not raw:
            return 'other'
        lower = raw.lower().strip()
        if lower in VALID_CATEGORIES:
            return lower
        # Try underscore variant
        underscored = lower.replace(' ', '_')
        if underscored in VALID_CATEGORIES:
            return underscored
        # Try known mappings
        if lower in CATEGORY_MAP:
            return CATEGORY_MAP[lower]
        return 'other'

    created = 0
    updated = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            code = safe_str(row['code'])
            if not code:
                continue

            name = safe_str(row['name'], '')
            name_ar = safe_str(row.get('name_ar'))
            category = normalize_category(safe_str(row['category']))
            unit = safe_str(row['unit'], 'EA')

            # Check if exists
            material = Material.query.filter_by(code=code).first()

            if material:
                # Update existing
                material.name = name
                material.name_ar = name_ar
                material.category = category
                material.unit = unit
                if 'current_stock' in row and pd.notna(row['current_stock']):
                    material.current_stock = float(row['current_stock'])
                if 'min_stock' in row and pd.notna(row['min_stock']):
                    material.min_stock = float(row['min_stock'])
                updated += 1
            else:
                # Create new
                material = Material(
                    code=code,
                    name=name,
                    name_ar=name_ar,
                    category=category,
                    unit=unit,
                    current_stock=float(row.get('current_stock', 0)) if pd.notna(row.get('current_stock')) else 0,
                    min_stock=float(row.get('min_stock', 0)) if pd.notna(row.get('min_stock')) else 0,
                    consumption_start_date=datetime.utcnow().date()
                )
                db.session.add(material)
                created += 1

        except Exception as e:
            db.session.rollback()
            errors.append(f"Row {idx + 2}: {str(e)}")

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ValidationError(f"Database error: {str(e)}")

    return jsonify({
        'status': 'success',
        'message': f'Import complete. Created: {created}, Updated: {updated}',
        'created': created,
        'updated': updated,
        'errors': errors
    }), 200


# ==================== HISTORY ====================

@bp.route('/<int:material_id>/history', methods=['GET'])
@jwt_required()
def get_stock_history(material_id):
    """Get stock change history."""
    from app.models import StockHistory
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    query = StockHistory.query.filter_by(material_id=material_id).order_by(StockHistory.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page)

    return jsonify({
        'status': 'success',
        'data': [h.to_dict() for h in pagination.items],
        'pagination': {'page': page, 'total': pagination.total}
    })


# ==================== BATCHES ====================

@bp.route('/<int:material_id>/batches', methods=['GET'])
@jwt_required()
def list_batches(material_id):
    """List batches for a material."""
    from app.models import MaterialBatch
    status = request.args.get('status')
    query = MaterialBatch.query.filter_by(material_id=material_id)
    if status:
        query = query.filter_by(status=status)
    batches = query.order_by(MaterialBatch.expiry_date.asc().nullslast()).all()
    return jsonify({'status': 'success', 'data': [b.to_dict() for b in batches]})


@bp.route('/batches/<int:batch_id>', methods=['GET'])
@jwt_required()
def get_batch(batch_id):
    """Get batch details."""
    from app.models import MaterialBatch
    batch = MaterialBatch.query.get_or_404(batch_id)
    return jsonify({'status': 'success', 'data': batch.to_dict()})


@bp.route('/batches/expiring', methods=['GET'])
@jwt_required()
def get_expiring_batches():
    """Get batches expiring soon."""
    days = request.args.get('days', 30, type=int)
    batches = stock_alerts.check_expiring_batches(days)
    return jsonify({'status': 'success', 'data': batches})


# ==================== LOCATIONS ====================

@bp.route('/locations', methods=['GET'])
@jwt_required()
def list_locations():
    """List storage locations."""
    from app.models import StorageLocation
    locations = StorageLocation.query.filter_by(is_active=True).all()
    return jsonify({'status': 'success', 'data': [l.to_dict() for l in locations]})


@bp.route('/locations', methods=['POST'])
@jwt_required()
def create_location():
    """Create storage location."""
    user = engineer_or_admin_required()
    data = request.get_json()
    from app.models import StorageLocation
    location = StorageLocation(
        code=data['code'],
        name=data['name'],
        name_ar=data.get('name_ar'),
        warehouse=data.get('warehouse'),
        zone=data.get('zone'),
        aisle=data.get('aisle'),
        shelf=data.get('shelf'),
        bin=data.get('bin')
    )
    db.session.add(location)
    db.session.commit()
    return jsonify({'status': 'success', 'data': location.to_dict()}), 201


@bp.route('/locations/<int:location_id>', methods=['PUT'])
@jwt_required()
def update_location(location_id):
    """Update storage location."""
    user = engineer_or_admin_required()
    from app.models import StorageLocation
    location = StorageLocation.query.get_or_404(location_id)
    data = request.get_json()
    for field in ['name', 'name_ar', 'warehouse', 'zone', 'aisle', 'shelf', 'bin', 'is_active']:
        if field in data:
            setattr(location, field, data[field])
    db.session.commit()
    return jsonify({'status': 'success', 'data': location.to_dict()})


# ==================== VENDORS ====================

@bp.route('/vendors', methods=['GET'])
@jwt_required()
def list_vendors():
    """List vendors."""
    from app.models import Vendor
    vendors = Vendor.query.filter_by(is_active=True).all()
    return jsonify({'status': 'success', 'data': [v.to_dict() for v in vendors]})


@bp.route('/vendors', methods=['POST'])
@jwt_required()
def create_vendor():
    """Create vendor."""
    user = engineer_or_admin_required()
    data = request.get_json()
    from app.models import Vendor
    vendor = Vendor(
        code=data['code'],
        name=data['name'],
        name_ar=data.get('name_ar'),
        contact_person=data.get('contact_person'),
        email=data.get('email'),
        phone=data.get('phone'),
        address=data.get('address'),
        payment_terms=data.get('payment_terms'),
        lead_time_days=data.get('lead_time_days')
    )
    db.session.add(vendor)
    db.session.commit()
    return jsonify({'status': 'success', 'data': vendor.to_dict()}), 201


@bp.route('/vendors/<int:vendor_id>', methods=['PUT'])
@jwt_required()
def update_vendor(vendor_id):
    """Update vendor."""
    user = engineer_or_admin_required()
    from app.models import Vendor
    vendor = Vendor.query.get_or_404(vendor_id)
    data = request.get_json()
    for field in ['name', 'name_ar', 'contact_person', 'email', 'phone', 'address', 'payment_terms', 'lead_time_days', 'is_active']:
        if field in data:
            setattr(vendor, field, data[field])
    db.session.commit()
    return jsonify({'status': 'success', 'data': vendor.to_dict()})


# ==================== ALERTS ====================

@bp.route('/alerts', methods=['GET'])
@jwt_required()
def get_alerts():
    """Get stock alert summary."""
    summary = stock_alerts.get_alert_summary()
    return jsonify({'status': 'success', 'data': summary})


@bp.route('/alerts/low-stock', methods=['GET'])
@jwt_required()
def get_low_stock_alerts():
    """Get low stock alerts."""
    alerts = stock_alerts.check_low_stock()
    return jsonify({'status': 'success', 'data': alerts})


@bp.route('/alerts/reorder', methods=['GET'])
@jwt_required()
def get_reorder_alerts():
    """Get reorder needed alerts."""
    alerts = stock_alerts.check_reorder_needed()
    return jsonify({'status': 'success', 'data': alerts})


# ==================== INVENTORY COUNT ====================

@bp.route('/counts', methods=['GET'])
@jwt_required()
def list_counts():
    """List inventory counts."""
    from app.models import InventoryCount
    counts = InventoryCount.query.order_by(InventoryCount.count_date.desc()).limit(20).all()
    return jsonify({'status': 'success', 'data': [c.to_dict() for c in counts]})


@bp.route('/counts', methods=['POST'])
@jwt_required()
def create_count():
    """Start a new inventory count."""
    user = engineer_or_admin_required()
    data = request.get_json()
    from app.models import InventoryCount
    count = InventoryCount(
        count_date=data.get('count_date', date.today()),
        count_type=data.get('count_type', 'full'),
        created_by_id=user.id
    )
    db.session.add(count)
    db.session.commit()
    return jsonify({'status': 'success', 'data': count.to_dict()}), 201


@bp.route('/counts/<int:count_id>/items', methods=['POST'])
@jwt_required()
def add_count_item(count_id):
    """Add/update count item."""
    user = get_current_user()
    data = request.get_json()
    from app.models import InventoryCount, InventoryCountItem, Material

    count = InventoryCount.query.get_or_404(count_id)
    material = Material.query.get_or_404(data['material_id'])

    item = InventoryCountItem.query.filter_by(
        count_id=count_id,
        material_id=data['material_id']
    ).first()

    if not item:
        item = InventoryCountItem(
            count_id=count_id,
            material_id=data['material_id'],
            system_quantity=material.current_stock
        )
        db.session.add(item)

    item.counted_quantity = data['counted_quantity']
    item.variance = item.counted_quantity - item.system_quantity
    item.counted_by_id = user.id
    item.counted_at = datetime.utcnow()
    item.notes = data.get('notes')

    db.session.commit()
    return jsonify({'status': 'success', 'data': item.to_dict()})


@bp.route('/counts/<int:count_id>/approve', methods=['POST'])
@jwt_required()
def approve_count(count_id):
    """Approve count and apply adjustments."""
    user = engineer_or_admin_required()
    from app.models import InventoryCount

    count = InventoryCount.query.get_or_404(count_id)
    count.status = 'approved'
    count.approved_by_id = user.id

    # Apply adjustments
    for item in count.items:
        if item.variance != 0:
            MaterialService.adjust(
                material_id=item.material_id,
                new_quantity=item.counted_quantity,
                user_id=user.id,
                reason=f'Inventory count adjustment (Count #{count_id})'
            )

    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Count approved and adjustments applied'})


# ==================== AI FEATURES ====================

@bp.route('/ai/reorder-prediction/<int:material_id>', methods=['GET'])
@jwt_required()
def predict_reorder(material_id):
    """Get AI reorder prediction."""
    prediction = material_ai.predict_reorder_date(material_id)
    return jsonify({'status': 'success', 'data': prediction})


@bp.route('/ai/demand-forecast/<int:material_id>', methods=['GET'])
@jwt_required()
def forecast_demand(material_id):
    """Get AI demand forecast."""
    days = request.args.get('days', 30, type=int)
    forecast = material_ai.forecast_demand(material_id, days)
    return jsonify({'status': 'success', 'data': forecast})


@bp.route('/ai/anomalies', methods=['GET'])
@jwt_required()
def get_consumption_anomalies():
    """Get consumption anomalies."""
    material_id = request.args.get('material_id', type=int)
    anomalies = material_ai.detect_consumption_anomalies(material_id)
    return jsonify({'status': 'success', 'data': anomalies})


@bp.route('/ai/optimal-reorder/<int:material_id>', methods=['GET'])
@jwt_required()
def get_optimal_reorder(material_id):
    """Get AI-calculated optimal reorder settings."""
    settings = material_ai.calculate_optimal_reorder_point(material_id)
    return jsonify({'status': 'success', 'data': settings})


@bp.route('/ai/cost-optimization', methods=['GET'])
@jwt_required()
def get_cost_optimization():
    """Get cost optimization suggestions."""
    suggestions = material_ai.suggest_cost_optimization()
    return jsonify({'status': 'success', 'data': suggestions})


@bp.route('/ai/insights', methods=['GET'])
@jwt_required()
def get_material_insights():
    """Get AI usage insights."""
    material_id = request.args.get('material_id', type=int)
    insights = material_ai.get_usage_insights(material_id)
    return jsonify({'status': 'success', 'data': insights})


@bp.route('/ai/search', methods=['POST'])
@jwt_required()
def natural_language_search():
    """Natural language material search."""
    data = request.get_json()
    result = material_ai.natural_language_search(data.get('query', ''))
    return jsonify({'status': 'success', 'data': result})


@bp.route('/ai/abc-analysis', methods=['GET'])
@jwt_required()
def get_abc_analysis():
    """Get ABC inventory analysis."""
    analysis = material_ai.get_abc_analysis()
    return jsonify({'status': 'success', 'data': analysis})


@bp.route('/ai/dead-stock', methods=['GET'])
@jwt_required()
def get_dead_stock():
    """Get dead stock report."""
    months = request.args.get('months', 6, type=int)
    dead_stock = material_ai.get_dead_stock(months)
    return jsonify({'status': 'success', 'data': dead_stock})


@bp.route('/ai/budget-forecast', methods=['GET'])
@jwt_required()
def get_budget_forecast():
    """Get budget forecast."""
    days = request.args.get('days', 30, type=int)
    forecast = material_ai.forecast_budget(days)
    return jsonify({'status': 'success', 'data': forecast})


# ==================== REPORTS ====================

@bp.route('/reports/consumption', methods=['GET'])
@jwt_required()
def get_consumption_report():
    """Get consumption report."""
    period = request.args.get('period', 'monthly')
    report = material_ai.generate_consumption_report(period)
    return jsonify({'status': 'success', 'data': report})


@bp.route('/reports/export', methods=['GET'])
@jwt_required()
def export_materials():
    """Export materials to Excel."""
    # Generate Excel export
    # Return file download
    pass


# ==================== ROUTE ALIASES (frontend compatibility) ====================

@bp.route('/analytics/abc', methods=['GET'])
@jwt_required()
def abc_analysis_alias():
    """Alias for /ai/abc-analysis - frontend uses /analytics/abc."""
    analysis = material_ai.get_abc_analysis()
    return jsonify({'status': 'success', 'data': analysis})


@bp.route('/analytics/dead-stock', methods=['GET'])
@jwt_required()
def dead_stock_alias():
    """Alias for /ai/dead-stock - frontend uses /analytics/dead-stock."""
    months = request.args.get('months', 6, type=int)
    dead_stock = material_ai.get_dead_stock(months)
    return jsonify({'status': 'success', 'data': dead_stock})


@bp.route('/analytics/budget-forecast', methods=['GET'])
@jwt_required()
def budget_forecast_alias():
    """Alias for /ai/budget-forecast - frontend uses /analytics/budget-forecast."""
    days = request.args.get('days', 30, type=int)
    forecast = material_ai.forecast_budget(days)
    return jsonify({'status': 'success', 'data': forecast})


# ==================== CONSUMPTION HISTORY ====================

class MaterialConsumptionHistory(db.Model):
    """Imported or system-recorded consumption history per period."""
    __tablename__ = 'material_consumption_history'

    id = db.Column(db.Integer, primary_key=True)
    material_id = db.Column(db.Integer, db.ForeignKey('materials.id', ondelete='CASCADE'), nullable=False, index=True)
    period_type = db.Column(db.String(20), nullable=False)   # monthly | quarterly | yearly
    period_label = db.Column(db.String(20), nullable=False)  # "2025-03" | "2025-Q1" | "2025"
    quantity_consumed = db.Column(db.Float, nullable=False, default=0)
    source = db.Column(db.String(20), nullable=False, default='imported')  # imported | system
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, server_default=db.text('NOW()'), nullable=False)

    material = db.relationship('Material', backref=db.backref('consumption_history', lazy='dynamic'))

    def to_dict(self):
        return {
            'id': self.id,
            'material_id': self.material_id,
            'period_type': self.period_type,
            'period_label': self.period_label,
            'quantity_consumed': self.quantity_consumed,
            'source': self.source,
            'notes': self.notes,
            'created_at': self.created_at.isoformat() if self.created_at else None,
        }


@bp.route('/<int:material_id>/consumption-history', methods=['GET'])
@jwt_required()
def get_consumption_history(material_id):
    """
    Get consumption history for a material.
    Query params: period=monthly|quarterly|yearly, years=1|2|3 (default 2)
    """
    material = db.session.get(Material, material_id)
    if not material:
        raise NotFoundError("Material not found")

    period = request.args.get('period', 'monthly')
    years = request.args.get('years', 2, type=int)

    records = MaterialConsumptionHistory.query.filter_by(
        material_id=material_id,
        period_type=period
    ).order_by(MaterialConsumptionHistory.period_label).all()

    return jsonify({
        'status': 'success',
        'data': [r.to_dict() for r in records],
        'material': {'id': material.id, 'code': material.code, 'name': material.name, 'unit': material.unit},
    })


@bp.route('/consumption-history', methods=['GET'])
@jwt_required()
def get_all_consumption_history():
    """
    Get consumption history for all materials aggregated by period.
    Query params: period=monthly|quarterly|yearly
    """
    period = request.args.get('period', 'monthly')

    from sqlalchemy import func as sqlfunc
    rows = db.session.query(
        MaterialConsumptionHistory.period_label,
        sqlfunc.sum(MaterialConsumptionHistory.quantity_consumed).label('total')
    ).filter_by(period_type=period).group_by(
        MaterialConsumptionHistory.period_label
    ).order_by(MaterialConsumptionHistory.period_label).all()

    return jsonify({
        'status': 'success',
        'data': [{'period_label': r.period_label, 'total': float(r.total)} for r in rows],
    })


@bp.route('/consumption-import', methods=['POST'])
@jwt_required()
def import_consumption_history():
    """
    Import historical consumption data from Excel.
    Supports two formats:
    - New: year columns (2020, 2021, ...) with optional monthly detail sheet
    - Legacy: code, period_type, period_label, quantity_consumed columns
    """
    import pandas as pd
    import re
    from io import BytesIO

    engineer_or_admin_required()

    if 'file' not in request.files:
        raise ValidationError("No file provided")

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise ValidationError("File must be Excel (.xlsx or .xls)")

    try:
        file_bytes = BytesIO(file.read())
        xls = pd.ExcelFile(file_bytes)
        df = pd.read_excel(xls, sheet_name=0)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        raise ValidationError(f"Failed to read Excel file: {e}")

    # Detect format: new (year columns) or legacy (period_type column)
    col_lower = [c.lower().replace(' ', '_') for c in df.columns]
    is_legacy = 'period_type' in col_lower and 'quantity_consumed' in col_lower

    MONTHS = [('Jan', '01'), ('Feb', '02'), ('Mar', '03'), ('Apr', '04'),
              ('May', '05'), ('Jun', '06'), ('Jul', '07'), ('Aug', '08'),
              ('Sep', '09'), ('Oct', '10'), ('Nov', '11'), ('Dec', '12')]

    imported, updated, errors = 0, 0, []

    if is_legacy:
        # ── Legacy format: code, period_type, period_label, quantity_consumed ──
        df.columns = col_lower
        for _, row in df.iterrows():
            try:
                code = str(row['code']).strip()
                if not code or code.lower() == 'nan':
                    continue
                material = Material.query.filter_by(code=code).first()
                if not material:
                    errors.append(f"Material not found: {code}")
                    continue

                period_type = str(row['period_type']).strip()
                period_label = str(row['period_label']).strip()
                qty = float(row['quantity_consumed'])
                notes = str(row.get('notes', '') or '').strip() or None
                if notes and notes.lower() == 'nan':
                    notes = None

                existing = MaterialConsumptionHistory.query.filter_by(
                    material_id=material.id, period_type=period_type, period_label=period_label,
                ).first()

                if existing:
                    existing.quantity_consumed = qty
                    existing.notes = notes
                    updated += 1
                else:
                    db.session.add(MaterialConsumptionHistory(
                        material_id=material.id, period_type=period_type,
                        period_label=period_label, quantity_consumed=qty,
                        source='imported', notes=notes,
                    ))
                    imported += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f"Row error ({row.get('code', '?')}): {e}")
    else:
        # ── New format: year columns auto-detected ──
        if 'code' not in col_lower:
            raise ValidationError("Missing required 'code' column")

        # Find year columns (4-digit numbers 2000-2099)
        year_cols = []
        for c in df.columns:
            try:
                y = int(c)
                if 2000 <= y <= 2099:
                    year_cols.append((c, str(y)))
            except (ValueError, TypeError):
                pass

        if not year_cols:
            raise ValidationError("No year columns found. Add columns like 2020, 2021, 2022, etc.")

        # Map column names to lowercase for code lookup
        code_col = next(c for c in df.columns if c.lower().strip() == 'code')

        # ── Sheet 1: Yearly totals ──
        for idx, row in df.iterrows():
            try:
                code = str(row[code_col]).strip()
                if not code or code.lower() == 'nan':
                    continue
                material = Material.query.filter_by(code=code).first()
                if not material:
                    errors.append(f"Row {idx + 2}: Material not found: {code}")
                    continue

                for col_name, year_str in year_cols:
                    val = row.get(col_name)
                    if pd.isna(val):
                        continue
                    try:
                        qty = float(val)
                    except (ValueError, TypeError):
                        continue
                    if qty <= 0:
                        continue

                    existing = MaterialConsumptionHistory.query.filter_by(
                        material_id=material.id, period_type='yearly', period_label=year_str,
                    ).first()

                    if existing:
                        existing.quantity_consumed = qty
                        updated += 1
                    else:
                        db.session.add(MaterialConsumptionHistory(
                            material_id=material.id, period_type='yearly',
                            period_label=year_str, quantity_consumed=qty, source='imported',
                        ))
                        imported += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f"Row {idx + 2} ({row.get(code_col, '?')}): {e}")

        # ── Sheet 2: Monthly detail (optional) ──
        monthly_sheet_names = [s for s in xls.sheet_names if 'monthly' in s.lower()]
        if monthly_sheet_names:
            try:
                df_monthly = pd.read_excel(xls, sheet_name=monthly_sheet_names[0])
                df_monthly.columns = [str(c).strip() for c in df_monthly.columns]
                m_code_col = next((c for c in df_monthly.columns if c.lower().strip() == 'code'), None)
                m_year_col = next((c for c in df_monthly.columns if c.lower().strip() == 'year'), None)

                if m_code_col and m_year_col:
                    for idx, row in df_monthly.iterrows():
                        try:
                            code = str(row[m_code_col]).strip()
                            if not code or code.lower() == 'nan':
                                continue
                            material = Material.query.filter_by(code=code).first()
                            if not material:
                                errors.append(f"Monthly row {idx + 2}: Material not found: {code}")
                                continue

                            year_val = str(int(row[m_year_col]))

                            for month_name, month_num in MONTHS:
                                # Try exact match and case-insensitive
                                val = None
                                for c in df_monthly.columns:
                                    if c.strip().lower()[:3] == month_name.lower():
                                        val = row.get(c)
                                        break
                                if val is None or pd.isna(val):
                                    continue
                                try:
                                    qty = float(val)
                                except (ValueError, TypeError):
                                    continue
                                if qty <= 0:
                                    continue

                                period_label = f"{year_val}-{month_num}"
                                existing = MaterialConsumptionHistory.query.filter_by(
                                    material_id=material.id, period_type='monthly',
                                    period_label=period_label,
                                ).first()

                                if existing:
                                    existing.quantity_consumed = qty
                                    updated += 1
                                else:
                                    db.session.add(MaterialConsumptionHistory(
                                        material_id=material.id, period_type='monthly',
                                        period_label=period_label, quantity_consumed=qty,
                                        source='imported',
                                    ))
                                    imported += 1
                        except Exception as e:
                            db.session.rollback()
                            errors.append(f"Monthly row {idx + 2}: {e}")
            except Exception as e:
                errors.append(f"Monthly sheet error: {e}")

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise ValidationError(f"Database error: {str(e)}")

    return jsonify({
        'status': 'success',
        'imported': imported,
        'updated': updated,
        'errors': errors,
    }), 200


@bp.route('/consumption-template', methods=['GET'])
def consumption_import_template():
    """Download Excel template for consumption history import, pre-populated with materials from DB."""
    import traceback
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from io import BytesIO
        from flask import send_file

        wb = openpyxl.Workbook()

        # Styling
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        year_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        # Get equipment types for dropdown
        try:
            from app.models import Equipment
            eq_types = sorted(set(
                t[0] for t in db.session.query(Equipment.equipment_type).distinct().all() if t[0]
            ))
        except Exception:
            eq_types = []

        # Get all active materials
        materials = Material.query.filter_by(is_active=True).order_by(Material.code).all()

        # Year range
        current_year = datetime.utcnow().year
        years = list(range(2020, current_year + 1))

        # ═══ Sheet 1: Yearly Summary ═══
        ws = wb.active
        ws.title = 'Yearly Summary'

        headers = ['code', 'name', 'unit'] + [str(y) for y in years] + ['equipment_type']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        year_start = 4  # column D
        year_end = year_start + len(years) - 1

        # Pre-populate with existing materials
        for row_idx, mat in enumerate(materials, 2):
            ws.cell(row=row_idx, column=1, value=mat.code)
            ws.cell(row=row_idx, column=2, value=mat.name)
            ws.cell(row=row_idx, column=3, value=mat.unit or 'EA')
            for col_idx in range(year_start, year_end + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = year_fill
                cell.number_format = '#,##0'

        # Equipment type dropdown
        if eq_types:
            eq_list = '"' + ','.join(eq_types[:250]) + '"'  # Excel limit ~255 chars
            if len(eq_list) > 255:
                # Too many types for inline list — use a reference sheet instead
                eq_list = '"' + ','.join(eq_types[:20]) + '"'
            eq_col_idx = year_end + 1
            dv = DataValidation(type='list', formula1=eq_list, allow_blank=True)
            dv.error = 'Please select a valid equipment type'
            ws.add_data_validation(dv)
            eq_col_letter = get_column_letter(eq_col_idx)
            last_row = max(len(materials) + 1, 2) + 100
            dv.add(f'{eq_col_letter}2:{eq_col_letter}{last_row}')

        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        for i in range(year_start, year_end + 1):
            ws.column_dimensions[get_column_letter(i)].width = 10
        ws.column_dimensions[get_column_letter(year_end + 1)].width = 20
        ws.freeze_panes = 'D2'

        # ═══ Sheet 2: Monthly Detail (optional) ═══
        ws2 = wb.create_sheet('Monthly Detail')
        month_headers = ['code', 'year', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        for col_idx, h in enumerate(month_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 8
        for i in range(3, 15):
            ws2.column_dimensions[get_column_letter(i)].width = 8
        ws2.freeze_panes = 'C2'

        # ═══ Sheet 3: Instructions ═══
        ws3 = wb.create_sheet('Instructions')
        instructions = [
            ['CONSUMPTION HISTORY IMPORT — INSTRUCTIONS'],
            [''],
            ['SHEET 1: Yearly Summary (required)'],
            ['code — Material code (must match system)'],
            ['name — Pre-filled for reference (not imported)'],
            ['unit — Pre-filled for reference'],
            ['2020, 2021, ... — Yearly consumption. Leave blank if no data.'],
            ['equipment_type — Optional. Select from dropdown.'],
            [''],
            ['SHEET 2: Monthly Detail (optional)'],
            ['code — Material code (must match system)'],
            ['year — e.g. 2024'],
            ['Jan-Dec — Monthly quantity. Leave blank if no data.'],
            [''],
            ['TIPS:'],
            ['• Add more year columns (2026, 2027...) — auto-detected on import'],
            ['• Empty cells are ignored'],
            ['• Existing data is updated, not duplicated'],
            ['• Monthly Detail sheet is optional'],
        ]
        if eq_types:
            instructions.append([''])
            instructions.append(['VALID EQUIPMENT TYPES:'])
            for et in eq_types:
                instructions.append([et])

        for row_data in instructions:
            ws3.append(row_data)

        ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws3.column_dimensions['A'].width = 60

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='consumption_history_template.xlsx'
        )
    except Exception as e:
        import logging
        logging.getLogger('app').error(f'Template generation failed: {traceback.format_exc()}')
        return jsonify({'status': 'error', 'message': f'Template generation failed: {str(e)}'}), 500
